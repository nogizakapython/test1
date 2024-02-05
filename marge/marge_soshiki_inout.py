#######   組織改編、データのマージ処理    ###################
#######   新規作成   2024/1/17         #####################
#######   作成者     takao.hattori     #####################
#######   クラスファイルの用途
#######   組織改編情報のデータを１つのファイルにマージする
#######
#######   クラスファイルの使い方        
#######   (1) 下記のコードを記載して、クラスファイルを読み込む
#######   　　from marge_soshiki_inout import Marge_soshiki_output
#######   (2)　インスタンスを読み込む
#######       <1> インスタンスを呼び出す
#######       r_marge = Marge_soshiki_output(work_file_list,export_file,startnum,soshiki_max_count)
#######       <2> データのマージ処理
#######       r_marge.main()
#######       
############################################################

# ライブラリの呼び出し
import openpyxl

class Marge_soshiki_output():
    #コンストラクタ
    def __init__(self,work_file_list,export_file,startnum,soshiki_max_count):
        self.work_file_list = work_file_list
        self.export_file = export_file
        self.startnum = startnum
        self.soshiki_max_count = soshiki_max_count

    # マージ対象ファイルを読み込んで作業者が転記したファイルを1つのファイルにマージする。
    def main(self): 
        #マージ先のシート名を定義する      
        sheet_name = "組織改編"  
        # 人事情報のマージ元一覧ファイルを読み込む 
        with open(self.work_file_list,mode="r",encoding="utf-8") as f2:
            while True:
                filename2 = f2.readline()
                filename2 = filename2.replace("\n","")
                # 作業用ファイル一覧ファイルのデータがNULLの時、ループを抜ける。
                if filename2 == '':
                    break
                
                #マージ先のシート名を定義する  
                wb = openpyxl.load_workbook(filename2)
                ws = wb[sheet_name]
                # マージ先の出力ファイルを定義する
                output1 = openpyxl.load_workbook(self.export_file)
                op = output1[sheet_name]
                # マージ対象の作業ファイルから転記したデータをマージ先ファイルにコピーする。
                for row in ws.iter_rows(min_row = self.startnum):
                    company_name = row[2].value
                    # 会社名がNULLの時にループを抜ける
                    if company_name is None:
                        break
                    # シートにデータが0件の時、マージ先にデータを出力しない
                    if company_name == '=C4':
                        continue
                    changes_day = row[3].value
                    detail = row[4].value
                    release_day = row[5].value
                    input_day = row[6].value
                    const_sites = row[7].value
                    const_url = row[8].value
                    appending = row[9].value
                    op.cell(row=self.soshiki_max_count,column=3).value = company_name
                    op.cell(row=self.soshiki_max_count,column=4).value = changes_day
                    op.cell(row=self.soshiki_max_count,column=5).value = detail
                    op.cell(row=self.soshiki_max_count,column=6).value = release_day
                    op.cell(row=self.soshiki_max_count,column=7).value = input_day
                    op.cell(row=self.soshiki_max_count,column=8).value = const_sites
                    op.cell(row=self.soshiki_max_count,column=9).value = const_url
                    op.cell(row=self.soshiki_max_count,column=9).hyperlink = const_url
                    op.cell(row=self.soshiki_max_count,column=10).value = appending
                    # 出力先の行数を1加算する。
                    self.soshiki_max_count += 1
                # マージ先エクセルファイルを保存する。    
                output1.save(self.export_file)