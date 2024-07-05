import openpyxl as op
import msoffcrypto

file_name = "test2.xlsx"
sh_name = "Sheet1"
wb = op.load_workbook(file_name)
ws = wb[sh_name]
start_num = 3
i = start_num

while True:
    w_name = ws.cell(i,2).value
    if w_name == None:
        break
    w_syozoku = ws.cell(i,3).value
    print(w_name)
    print(w_syozoku)
    i += 1


