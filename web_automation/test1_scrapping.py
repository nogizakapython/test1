import asyncio
import os
from datetime import datetime

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# 設定
# ============================================================
BASE_URL = "http://localhost:5000/public/news"
OUTPUT_DIR = "output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "news_data.xlsx")


# ============================================================
# お知らせデータ取得
# ============================================================
async def fetch_news(page) -> list[dict]:
    """
    お知らせ一覧ページから全件のデータを取得する。
    構造:
      .news-list > .news-item
        .news-title    … タイトル
        .news-category … カテゴリ
        .news-date     … 日付
        .news-body     … 本文
    """
    print(f"[INFO] アクセス中: {BASE_URL}")
    await page.goto(BASE_URL, wait_until="networkidle")

    # お知らせカードを全件取得
    items = await page.query_selector_all(".news-list .news-item")
    print(f"[INFO] お知らせを {len(items)} 件検出しました")

    news_list = []
    for item in items:
        title_el = await item.query_selector(".news-title")
        cat_el   = await item.query_selector(".news-category")
        date_el  = await item.query_selector(".news-date")
        body_el  = await item.query_selector(".news-body")

        news_list.append({
            "title":    (await title_el.inner_text()).strip() if title_el else "",
            "category": (await cat_el.inner_text()).strip()   if cat_el  else "",
            "date":     (await date_el.inner_text()).strip()  if date_el else "",
            "body":     (await body_el.inner_text()).strip()  if body_el else "",
        })

    return news_list


# ============================================================
# Excel 保存
# ============================================================
def save_to_excel(news_list: list[dict], filepath: str) -> None:
    """
    取得したお知らせデータを Excel ファイルに保存する。
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "お知らせ一覧"

    # ── ヘッダー設定 ──────────────────────────
    headers = ["No.", "タイトル", "カテゴリ", "日付", "本文"]
    header_fill  = PatternFill(fill_type="solid", fgColor="6C757D")   # ヘッダーと同じグレー
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align

    # ── 列幅設定 ──────────────────────────────
    col_widths = {1: 6, 2: 36, 3: 12, 4: 14, 5: 70}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── データ書き込み ────────────────────────
    even_fill = PatternFill(fill_type="solid", fgColor="E8F0FE")   # カテゴリバッジ風の薄青

    for row_idx, item in enumerate(news_list, start=2):
        row_data = [
            row_idx - 1,           # No.
            item["title"],
            item["category"],
            item["date"],
            item["body"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align if col_idx in (1, 3, 4) else left_align
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # ── ヘッダー行を固定 ──────────────────────
    ws.freeze_panes = "A2"

    # ── メタ情報シート ────────────────────────
    ws_meta = wb.create_sheet(title="取得情報")
    ws_meta.append(["項目", "値"])
    ws_meta.append(["取得URL",     BASE_URL])
    ws_meta.append(["取得件数",    len(news_list)])
    ws_meta.append(["取得日時",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_meta.append(["出力ファイル", filepath])
    ws_meta.column_dimensions["A"].width = 16
    ws_meta.column_dimensions["B"].width = 50

    wb.save(filepath)
    print(f"[INFO] Excel 保存完了: {filepath}  ({len(news_list)} 件)")


# ============================================================
# メイン処理
# ============================================================
async def main() -> None:
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context()
        page    = await context.new_page()

        try:
            # ① お知らせを全件取得
            news_list = await fetch_news(page)

            if not news_list:
                print("[WARN] お知らせが0件でした。ページ内容を確認してください。")
                return

            # ② 取得内容をコンソールに表示
            print("-" * 60)
            for i, item in enumerate(news_list, start=1):
                print(f"[{i}] {item['date']} 【{item['category']}】 {item['title']}")
            print("-" * 60)

            # ③ Excel 保存
            save_to_excel(news_list, OUTPUT_FILE)

        except Exception as e:
            print(f"[ERROR] 処理中にエラーが発生しました: {e}")
            raise
        finally:
            await context.close()
            await browser.close()


if __name__ == "__main__":
    asyncio.run(main())