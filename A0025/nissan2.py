#######   日産自動車 中計・決算ニュース情報取得ツール2　###########
#######   新規作成  2024/07/24  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
# date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "nissan2" + date1 + ".txt"
out_file = "nissan2.txt"
date_str = ""
# 2026/4/22 ニュースリリースの詳細ページのドメイン変更に伴う修正
base_url = 'https://global.nissannews.com/'
web_url = 'https://global.nissannews.com/ja-JP/channels/news?selectedTabId=news-releases'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
xpath_str2 = ""
w_url = ""
t_year = 0
year_array=[]
w_titlehead = ""

# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
target_url = web_url

try:
   driver.get(target_url)
   sleep(3)
   for i in range(1,1001):

         # 2026/4/22 takao.hattori xpath変更に伴う修正
         try:
            xpath_str1 = '//*[@id="releases-list releases-content"]/div/article[' + str(i) + ']/div/div[1]/date/time'
            
         except:
             break   
         xpath_str2 = '//*[@id="releases-list releases-content"]/div/article[' + str(i) + ']/div/div[2]/a'
         element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
         print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
         element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
         print(element_str2.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
        
          
      
       
except EnvironmentError as e:
   str100 = e     
except:
   str100 = 1
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
     w_urlstr = ""
     w_titlestr = ""
     w_ymdstr = ""
     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     if line1:
       row_count += 1
     else:
       break 
     #2026/4/22 検索条件変更に伴う修正  
     result1 = re.search('          2',line1)
     result2 = re.match('<a href',line1)
     result3 = re.search('      日産',line1)
     
     #2026/4/22 検索条件変更に伴うURL、掲載日時、タイトルの取得処理の修正
     if result1:
           w_ymdstr = line1
           w_ymd = w_ymdstr.replace(' ','')
         #   print(w_ymd)
     if result2:
           url_array = line1.split("=")
           w_urlstr = url_array[1]
           w_urlstr = w_urlstr.replace('>','')
           w_urlstr = w_urlstr.replace('"','')
           result4 = re.match('https',w_urlstr)
           if result4:
              w_url = w_urlstr
           else:
              w_url = base_url + w_urlstr   
         #   print(w_url)
     if result3:
           w_titlestr = line1
           w_title = w_titlestr.replace(" ",'')
         #   print(w_title) 
            # 2026/4/22 検索条件の修正
           keyword = r"(中期経営計画|報告書|レポート|経営|経営計画|決算)"
           title_result1 = re.search(keyword,w_title)
           if title_result1:
  
              wb = op.load_workbook(export_file)
              sh_name = '日産自動車2'
              ws = wb[sh_name]
              ws.cell(row=max_row,column=2).value = w_title
              ws.cell(row=max_row,column=3).value = w_url
              ws.cell(row=max_row,column=4).value = w_ymd
              ws.cell(row=max_row,column=6).value = w_url
              ws.cell(row=max_row,column=6).hyperlink = w_url
              ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
              max_row += 1
              # エクセルファイルの保存
              wb.save(export_file)
