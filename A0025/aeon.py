#######   イオン 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/19  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
w_ymd = dt.strftime('%Y年%m年%d日')


input_file = "aeon" + date1 + ".txt"
out_file = "aeon.txt"
date_str = ""
w_title = ""
base_url = 'https://www.aeon.info'

target_url = 'https://www.aeon.info/ir/news/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
# Chromeを指定する

driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
try:
    driver.get(target_url)
    sleep(5)
    

    
    for i in range(1,51):
        
        
        try:
            xpath_str1 = '//*[@id="wrap-container"]/main/div[2]/div[2]/div/div/div/dl/dt[' + str(i) + ']'
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
            print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
            xpath_str2 = '//*[@id="wrap-container"]/main/div[2]/div[2]/div/div/div/dl/dd[' + str(i) + ']'
            element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
            print(element_str2.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))    
        except:
            print(1)
            break
        

except EnvironmentError as e:
    str100 = e     
except:
    str1000 = ""

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result1 = re.search('<span class="eirItem_date date-time">',line1)
    result2 = re.search('<a href',line1)
    

    if result1:
       w_array1 = line1.split(">")
       w_ymdstr = w_array1[1]
       w_ymd = w_ymdstr.replace("</span","")
    #    print(w_ymd)

    if result2:   
       w_array2 = line1.split(">")
       w_urlstr = w_array2[0]
       url_array = w_urlstr.split("=")
       w_urlstr = url_array[1]
       w_urlstr = w_urlstr.replace(" target",'')
       w_url = w_urlstr.replace('"','')
    #    print(w_url)
       
       w_titlestr = w_array2[1]
       w_title = w_titlestr.replace('<!--','')
    #    print(w_title)
       key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
       title_result = re.search(key_word,w_title)
       if title_result:
            wb = op.load_workbook(export_file)
            sh_name = 'AEON'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
            ws.cell(row=max_row,column=6).hyperlink = w_url
            ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)
          

