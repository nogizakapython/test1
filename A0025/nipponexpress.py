#######   日本通運 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font
from selenium.webdriver.support.ui import Select


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "nipponexpress" + date1 + ".txt"
out_file = "nipponexpress.txt"
date_str = ""
w_title = ""
base_url = 'https://www.nipponexpress-holdings.com'
web_url = 'https://www.nipponexpress-holdings.com/ja/ir/library/news/archive.html?year='
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
t_year = 0
year_array=[]

# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする



for year in [date5,date6]:
    target_url = web_url + str(year)
    try:
       driver.get(target_url)
       sleep(3)

       for i in range(1,26):
          try:
             xpath_str1 = '//*[@id="irp-press-list-archive"]/div[' + str(i) + ']'
             
                
          except:
             break    
          element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
          print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
      
       
    except EnvironmentError as e:
       str100 = e     
    except:
       print("データ件数0件です",file=codecs.open(input_file,'a','utf-8'))
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    w_urlstr = ""
    w_titlestr = ""
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
      row_count += 1
    else:
      break   

    result1 = re.match('<div class=',line1)

    if result1:
       w_array1 = line1.split(">")
       w_ymdstr = w_array1[3]
       w_ymd = w_ymdstr.replace('</div','')
      #  print(w_ymd)    

       w_urlstr = w_array1[7]
       url_array = w_urlstr.split("=")
       w_urlstr = url_array[1]
       w_urlstr = w_urlstr.replace('target','')
       w_urlstr = w_urlstr.replace('"','')
       url_result = re.match('https',w_urlstr)
       if url_result:
          w_url = w_urlstr
       else:
          w_url = 'https:' + w_urlstr   
      #  print(w_url)

       w_titlestr = w_array1[8]
       w_titlestr = w_titlestr.replace('</a','')
       w_title = w_titlestr.replace('<span class="irp-fsize"','')
      #  print(w_title)    


       key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
       title_result = re.search(key_word,w_title)
       if title_result:
          wb = op.load_workbook(export_file)
          sh_name = '日本通運'
          ws = wb[sh_name]
          ws.cell(row=max_row,column=2).value = w_title
          ws.cell(row=max_row,column=3).value = w_url
          ws.cell(row=max_row,column=4).value = w_ymd
          ws.cell(row=max_row,column=6).value = w_url
          ws.cell(row=max_row,column=6).hyperlink = w_url
          ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
          max_row += 1
          # エクセルファイルの保存
          wb.save(export_file)
