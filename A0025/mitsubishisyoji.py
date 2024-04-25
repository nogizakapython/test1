#######   三菱商事 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font
from selenium.webdriver.support.ui import Select


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "mitsubishisyoji" + date1 + ".txt"
out_file = "mitsubishisyoji.txt"
date_str = ""
w_title = ""
base_url1 = 'https://www.mitsubishicorp.com/jp/ja'
base_url2 = 'https://www.mitsubishicorp.com'
web_url = 'https://www.mitsubishicorp.com/jp/ja/ir/latest/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
w_year = ""
year_array=[]

# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
target_url = web_url
try:
   driver.get(target_url)
   sleep(3)

   for i in range(1,3):
      try:
         
         xpath_str1 = '//*[@id="C_Content"]/section/section[' + str(i) + ']'
      except:
         break
      element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
      print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
      
       
except EnvironmentError as e:
    str100 = e     
except:
    str100 = "a"
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
   w_urlstr = ""
   w_titlestr = ""
   line1 = fileobj.readline()
   line1 = line1.replace("\n","")
   if line1:
      row_count += 1
   elif row_count <= 650:
      row_count += 1  
   else:
     break   

   result1 = re.match('			<h2',line1)
   result2 = re.search('<th>',line1) 
   result3 = re.search('<td><a class',line1) 
   
   if result1:
      w_array1  = line1.split(">")
      w_yearstr = w_array1[1]
      w_year = w_yearstr.replace('</h2','')
      # print(w_year)

   if result2:
      w_array2 = line1.split(">")
      w_ymdstr = w_array2[1]
      w_ymdstr = w_ymdstr.replace('</th','')
      w_ymdstr = w_ymdstr.replace('月','/')
      w_ymdstr = w_ymdstr.replace('日','')
      ymd_array = w_ymdstr.split("/")
      month1 = int(ymd_array[0])
      day1 = int(ymd_array[1])
      if month1 < 10:
         month1 = "0" + str(month1)
      else:
         month1 = str(month1)
      if day1 < 10:
         day1 = "0" + str(day1)
      else:
         day1 = str(day1)
      w_ymd = w_year + "/" + month1 + "/" + day1 
      w_ymd = w_ymd.replace("年","")     

      
      # print(w_ymd)    

   if result3:
      w_array3 = line1.split(">")
      w_urllink1 = w_array3[1]
      w_urllink2 = w_array3[3]
      urllink_result = re.match('<a href',w_urllink2)
      if urllink_result:
         w_urlstr = w_urllink2
         w_titlestr = w_array3[4]
      else:
         w_urlstr = w_urllink1
         w_titlestr = w_array3[2]  
      
      w_urlstr = w_urlstr.replace('<a class="link" href=','')
      w_urlstr = w_urlstr.replace('<a href=','')
      w_urlstr = w_urlstr.replace('"','')
      urlresult1 = re.match('../',w_urlstr)
      urlresult2 = re.match('https',w_urlstr)
      if urlresult1:
         w_urlstr = w_urlstr.replace('../..','')
         w_url = base_url1 + w_urlstr
      elif urlresult2:
         w_url = w_urlstr
      else:
         w_url = base_url2 + w_urlstr
      # print(w_url)
         
      w_titlestr = w_titlestr.replace('</a','')
      w_title = w_titlestr.replace('<span class="tse"','')
      # print(w_title)

      key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
      title_result = re.search(key_word,w_title)
      if title_result:
          wb = op.load_workbook(export_file)
          sh_name = '三菱商事'
          ws = wb[sh_name]
          ws.cell(row=max_row,column=2).value = w_title
          ws.cell(row=max_row,column=3).value = w_url
          ws.cell(row=max_row,column=4).value = w_ymd
          ws.cell(row=max_row,column=6).value = w_url
          ws.cell(row=max_row,column=6).hyperlink = w_url
          ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
          max_row += 1
          # エクセルファイルの保存
          wb.save(export_file)
