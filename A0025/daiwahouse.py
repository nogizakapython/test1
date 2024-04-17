#######   大和ハウス 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/25  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')


input_file = "daiwahouse" + date1 + ".txt"
out_file = "daiwahouse.txt"
date_str = ""
w_title = ""
base_url1 = 'https://www.daiwahouse.co.jp'
web_url = 'https://www.daiwahouse.co.jp/ir/kessan/s_index.html'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""


# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
    
target_url = web_url
try:
    driver.get(target_url)
    sleep(5)
    for j in range(1,51):
        try:
            xpath_str1 = '//*[@id="main"]/div[2]/div/ul/li[' + str(j) + ']'
            
        except:
            break    
        element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
      
       
except EnvironmentError as e:
    str100 = e     
except:
    str100 = ""
     

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
       row_count += 1
    else:
       break   

    result1 = re.match('<a href',line1)
    

    if result1:
       w_array1 = line1.split("=")
       w_url = w_array1[1]
       w_url = w_url.replace(' target',"")
       w_url = w_url.replace('"',"")
    #    print(w_url)

       w_array2 = line1.split(">")
       w_ymdstr = w_array2[1]
       w_array3 = w_ymdstr.split("　")
       w_ymd = w_array3[0]
    #    print(w_ymd)
       title_array_len = len(w_array3)
       if title_array_len == 3:
            w_title = w_array3[1] + " " + w_array3[2]
       else:         
            w_title = w_array3[1]
       w_title = w_title.replace('</a','')
    #    print(w_title)
    
       key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
       title_result = re.search(key_word,w_title)
       if title_result:
           wb = op.load_workbook(export_file)
           sh_name = 'DAIWA'
           ws = wb[sh_name]
           ws.cell(row=max_row,column=2).value = w_title
           ws.cell(row=max_row,column=3).value = w_url
           ws.cell(row=max_row,column=4).value = w_ymd
           ws.cell(row=max_row,column=6).value = w_url
           ws.cell(row=max_row,column=6).hyperlink = w_url
           ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
           max_row += 1
           # エクセルファイルの保存
           wb.save(export_file)

