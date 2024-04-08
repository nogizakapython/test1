#######   asahi 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/19  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys 
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
w_ymd = dt.strftime('%Y年%m年%d日')


input_file = "asahi" + date1 + ".txt"
out_file = "asahi.txt"
date_str = ""
w_title = ""
base_url = 'https://www.asahigroup-holdings.com'

target_url = 'https://www.asahigroup-holdings.com/ir/news/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""


# エクセル出力関数 
def excel_output(w_title,w_url,w_ymd,max_row):
      wb = op.load_workbook(export_file)
      sh_name = 'ASAHI'
      ws = wb[sh_name]
      ws.cell(row=max_row,column=2).value = w_title
      ws.cell(row=max_row,column=3).value = w_url
      ws.cell(row=max_row,column=4).value = w_ymd
      ws.cell(row=max_row,column=6).value = w_url
      ws.cell(row=max_row,column=6).hyperlink = w_url
      ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
      # エクセルファイルの保存
      wb.save(export_file)



# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
try:
    driver.get(target_url)
    sleep(5)
    

    
    for i in range(1,16):
               
        try:
            xpath_str1 = '//*[@id="pagetop"]/main/div/ul/li[' + str(i) + ']/a'
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
            print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
        except:
            break
        
except EnvironmentError as e:
    str100 = e     
except:
    str1000 = ""

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result1 = re.match("<a href",line1)
    
    if result1:
        w_array1 = line1.split(">")
        w_line = w_array1[0]
        url_array = w_line.split("=")
        w_urlstr = url_array[1]
        w_urlstr = w_urlstr.replace('class','')
        w_urlstr = w_urlstr.replace('target','')
        w_url = w_urlstr.replace('"','')
        # print(w_url)

        w_ymdstr = w_array1[2]        
        w_ymdstr = w_ymdstr.replace('</time','')
        w_ymd = w_ymdstr.replace('.','/')
        # print(w_ymd)

        w_titlestr = w_array1[7]
        w_title = w_titlestr.replace('</span','')
        # print(w_title)

        key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
        title_result = re.search(key_word,w_title)
        if title_result:
          excel_output(w_title,w_url,w_ymd,max_row)
          max_row += 1
