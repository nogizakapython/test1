#######   MISUMI 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font
from selenium.webdriver.support.ui import Select


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "misumi" + date1 + ".txt"
out_file = "misumi.txt"
date_str = ""
w_title = ""
base_url = 'https://www.misumi.co.jp'
web_url = 'https://www.misumi.co.jp/ir/news/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
t_year = 0
year_array=[]

# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする

day_array1 = date7.split("/")
d_year = int(day_array1[0])
d_month = int(day_array1[1])
d_day = int(day_array1[2])

if d_month < 4:
    year_array.append(date6)
elif d_month == 4:
    year_array.append(date5)
    year_array.append(date6)
else:
    year_array.append(date5)

for year in year_array:
    target_url = web_url + str(year) + ".html"
    try:
       driver.get(target_url)
       sleep(3)

       for i in range(1,31):
          try:
             xpath_str1 = '//*[@id="main-content"]/section/div[3]/ul/li[' + str(i) + ']'
                  
          except:
             if i == 1:
                print("データ件数0件です",file=codecs.open(input_file,'a','utf-8'))      
             break    
          element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
          print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
      
       
    except EnvironmentError as e:
       str100 = e     
    except:
       print("データ件数0件です",file=codecs.open(input_file,'a','utf-8'))
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
   w_urlstr = ""
   w_titlestr = ""
   line1 = fileobj.readline()
   line1 = line1.replace("\n","")
   if line1:
     row_count += 1
   else:
     break   

   result1 = re.search('el-news__container',line1)
   result2 = re.search('el-news__date',line1) 
   result3 = re.match('        <p>',line1) 
   
   if result1:
      w_array1  = line1.split("=")
      w_urlstr = w_array1[2]
      w_urlstr = w_urlstr.replace('target','')
      w_urlstr = w_urlstr.replace('"','')
      w_url = base_url + w_urlstr 
      # print(w_url)

   if result2:
      w_array2 = line1.split("=")
      w_ymdstr = w_array2[2]
      w_ymdstr = w_ymdstr.replace('>','')
      w_ymdstr = w_ymdstr.replace('"','')
      w_ymd = w_ymdstr.replace('-','/')
      # print(w_ymd)    

   if result3:
      w_array3 = line1.split(">")
      w_titlestr = w_array3[1]
      w_title = w_titlestr.replace('</p','')
      # print(w_title)

      key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
      title_result = re.search(key_word,w_title)
      if title_result:
         wb = op.load_workbook(export_file)
         sh_name = 'ミスミグループ本社'
         ws = wb[sh_name]
         ws.cell(row=max_row,column=2).value = w_title
         ws.cell(row=max_row,column=3).value = w_url
         ws.cell(row=max_row,column=4).value = w_ymd
         ws.cell(row=max_row,column=6).value = w_url
         ws.cell(row=max_row,column=6).hyperlink = w_url
         ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
         max_row += 1
         # エクセルファイルの保存
         wb.save(export_file)
