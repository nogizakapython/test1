#######   日産自動車 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/04/01  ##########
#######   修正      2024/7/24  URLの変更とデザイン変更による改修
#######   修正      2024/7/29  xpath変更に伴う修正
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
# date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "nissan" + date1 + ".txt"
out_file = "nissan.txt"
date_str = ""
w_title = ""
base_url = 'https://www.nissan-global.com'
# URL修正 (2024/7/24)
web_url = 'https://www.nissan-global.com/JP/IR/NEWS_EVENT/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
t_year = 0
year_array=[]
w_titlehead = ""

# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
target_url = web_url

try:
   driver.get(target_url)
   sleep(3)
   for i in range(1,31):
      try:
            # xpathの修正(2024/7/25)
            # xpathの修正(2024/7/29)
            xpath_str1 = '/html/body/div[1]/div[2]/div/div/div/div/div[3]/div/div[1]/div/div[3]/div/div/div/div[2]/table/tbody/tr[' + str(i) + ']'
            
            
      except:
            break    
      element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
      print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
          
       
except EnvironmentError as e:
   str100 = e     
except:
   str100 = 1
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
     w_urlstr = ""
     w_titlestr = ""
     
     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     if line1:
       row_count += 1
     else:
       break   
   #   抽出タグの修正(2024/7/25)
     result0 = re.search('<th>',line1)
     result1 = re.match('							<td>',line1)
     result2 = re.match('" target="_blank">',line1)

     if result0:
         w_array1 = line1.split('>')
         w_ymd = w_array1[1]
         ymd_array = w_ymd.split('/')
         year1 = ymd_array[0]
         month1 = int(ymd_array[1])
         day_str = ymd_array[2]
         day_str = day_str.replace('<','')
         day1 = day_str.replace('<','')
         day1 = int(day_str)
         if month1 < 10:
            month1 = "0" + str(month1)
         else:
            month1 = str(month1)
         if day1 < 10:
            day1 = "0" + str(day1)
         else:
            day1 = str(day1)   

             
         w_ymd = year1 + "/" + month1 + "/" + day1
         # print(w_ymd)
        
     
     if result1:
         w_array2 = line1.split('>')
         w_urlstr = w_array2[1]
         url_array = w_urlstr.split('=')
         w_urlstr = url_array[1]
         w_urlstr = w_urlstr.replace(' target','')
         w_urlstr = w_urlstr.replace('"','')
         url_result = re.match('https',w_urlstr)
         next_row_flag = 0
         if url_result:
            w_url = w_urlstr
         else:   
            w_url = base_url + w_urlstr
          # print(w_url)
         try:
            w_titlestr = w_array2[2]
         except IndexError:
            next_row_flag = 1

         if next_row_flag == 0:
            w_title = w_titlestr.replace('</a','')
            # print(w_title)   
              
            keyword = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート|経営|経営計画)"
            title_result1 = re.search(keyword,w_title)
            if title_result1:
  
               wb = op.load_workbook(export_file)
               sh_name = '日産自動車'
               ws = wb[sh_name]
               ws.cell(row=max_row,column=2).value = w_title
               ws.cell(row=max_row,column=3).value = w_url
               ws.cell(row=max_row,column=4).value = w_ymd
               ws.cell(row=max_row,column=6).value = w_url
               ws.cell(row=max_row,column=6).hyperlink = w_url
               ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
               max_row += 1
               # エクセルファイルの保存
               wb.save(export_file)
     if result2:
         next_row_flag = 0
         w_array3 = line1.split('>')
         w_titlestr = w_array3[1]
         w_title = w_titlestr.replace('</a','')
         keyword = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート|経営|経営計画)"
         title_result1 = re.search(keyword,w_title)
         if title_result1:
  
               wb = op.load_workbook(export_file)
               sh_name = '日産自動車'
               ws = wb[sh_name]
               ws.cell(row=max_row,column=2).value = w_title
               ws.cell(row=max_row,column=3).value = w_url
               ws.cell(row=max_row,column=4).value = w_ymd
               ws.cell(row=max_row,column=6).value = w_url
               ws.cell(row=max_row,column=6).hyperlink = w_url
               ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
               max_row += 1
               # エクセルファイルの保存
               wb.save(export_file)