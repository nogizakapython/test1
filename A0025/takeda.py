#######   武田薬品工業 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/04/04  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "takeda" + date1 + ".txt"
out_file = "takeda.txt"
date_str = ""
w_title = ""
base_url = 'https://www.takeda.com'
web_url = 'https://www.takeda.com/jp/investors/financial-results/quarterly-results/'


max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
t_year = 0
year_array=[]
# Chromeを指定する
driver = webdriver.Chrome()


def title_cransing(str1):
   str1 = str1.replace("</span","")
   return str1

def url_cransing(str2):
   str2 = str2.replace('"',"")
   return str2

def excel_output_file(w_title,w_url,max_row):
   wb = op.load_workbook(export_file)
   sh_name = '武田薬品工業'
   ws = wb[sh_name]
   ws.cell(row=max_row,column=2).value = w_title
   ws.cell(row=max_row,column=3).value = w_url
#  ws.cell(row=max_row,column=4).value = w_ymd
   ws.cell(row=max_row,column=6).value = w_url
   ws.cell(row=max_row,column=6).hyperlink = w_url
   ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
   max_row += 1
   # エクセルファイルの保存
   wb.save(export_file)

# Chromeを開いて企業HPにアクセスする
target_url = web_url
   
try:
   driver.get(target_url)
   sleep(3)
   for i in range(1,15):
      try:
         xpath_str1 = '//*[@id="root"]/main/div[2]/section[' + str(i) + ']'
         element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
         print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
         
      except:
         if i < 10:
            continue
         else:
            break
          
       
except EnvironmentError as e:
   str100 = e     
except:
   str100 = 1
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
     w_urlstr = ""
     w_titlestr = ""
     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     row_count += 1
     

     result1 = re.search('data-section-name="## ',line1)
     result2 = re.search('data-section-name="###',line1)
     result3 = re.search('f_jp.pdf',line1)
     

     if result1:
        w_array1 = line1.split("=")
        w_titleheadstr = w_array1[3]
        titlehead_result = re.search("アーカイブ",w_titleheadstr)
        if titlehead_result:
           break
        else:
           w_titleheadstr = w_titleheadstr.replace('"##',"")
      
      #   print(w_titleheadstr)

     if result2:
        w_array2 = line1.split("=")
        w_titlemainstr = w_array2[3]
        w_titlemainstr = w_titlemainstr.replace('### ','')
        w_titlemainstr = w_titlemainstr.replace(' data-exclude-from-nav','')
        w_titlemainstr = w_titlemainstr.replace('"','')
      #   print(w_titlemainstr)


     if result3:
        patturn_str = r"(疫学情報|臨床試験)"
        archive_result = re.search(patturn_str,line1)
        if archive_result:
           continue
        else:
           w_array3 = line1.split(">")
           array_count = len(w_array3)
           
           if array_count == 105:
              w_titlefootstr1 = w_array3[8]
              w_titlefootstr1 = title_cransing(w_titlefootstr1)
              
              w_urlstr1 = w_array3[5]
              url_array1 = w_urlstr1.split('=')
              w_urlstr1 = url_array1[6]
              w_url1 = url_cransing(w_urlstr1)
              

              w_titlefootstr2 = w_array3[24]
              w_titlefootstr2 = title_cransing(w_titlefootstr2)
              w_urlstr2 = w_array3[21]
              
              url_array2 = w_urlstr2.split('=')
              w_urlstr2 = url_array2[6]
              w_url2 = url_cransing(w_urlstr2)
               
           else:
              w_titlefootstr1 = w_array3[14]
              w_titlefootstr1 = title_cransing(w_titlefootstr1)
              w_urlstr1 = w_array3[11]
              
              url_array1 = w_urlstr1.split('=')
              w_urlstr1 = url_array1[6]
              w_url1 = url_cransing(w_urlstr1)   
              
              w_titlefootstr2 = w_array3[30]
              w_titlefootstr2 = title_cransing(w_titlefootstr2)
              
              w_urlstr2 = w_array3[27]
              url_array2 = w_urlstr2.split('=')
              w_urlstr2 = url_array2[6]
              w_url2 = url_cransing(w_urlstr2)
              
            
           w_title1 = w_titleheadstr + " " + w_titlemainstr + " " + w_titlefootstr1
           excel_output_file(w_title1,w_url1,max_row)
           max_row += 1
           
           w_title2 = w_titleheadstr + " " + w_titlemainstr + " " + w_titlefootstr2
           excel_output_file(w_title2,w_url2,max_row)
           max_row += 1
