#######   矢崎総業 中計、決算ニュース情報取得ツール　###########
#######   新規作成  2024/04/08  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')


input_file = "yazakisogyo" + date1 + ".txt"
out_file = "yazakisogyo.txt"
date_str = ""
w_title = ""
base_url = 'https://catr.jp'
web_url = 'https://catr.jp/search?utf8=%E2%9C%93&word=%E7%9F%A2%E5%B4%8E%E7%B7%8F%E6%A5%AD%E6%A0%AA%E5%BC%8F%E4%BC%9A%E7%A4%BE&commit=%E6%A4%9C%E7%B4%A2'
target_url = 'https://catr.jp/companies/36b5f/32204'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""


# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
    
try:
    driver.get(target_url)
    sleep(5)
    for i in range(1,4):
      try:
         xpath_str1 = '/html/body/div/div[5]/div[1]/div[2]/table/tbody/tr[' + str(i) + ']'
         
      except:
         msg = "データがありません"

      element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
      print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
      
       
except EnvironmentError as e:
    str100 = e     
except:
    str100 = ""
     

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
       row_count += 1
    else:
       break   
    result1 = re.search('<td><a href',line1)
            
    if result1:
       w_array1 = line1.split(">")
       w_urlstr = w_array1[1]
       w_urlstr = w_urlstr.replace('<a href=',"")
       w_urlstr = w_urlstr.replace('"',"")
       w_url = base_url + w_urlstr
       
       w_title = w_array1[2]
       w_title = w_title.replace('</a','')
       
       w_ymd = w_title
       w_ymd = w_ymd.replace("年",'/')
       w_ymd = w_ymd.replace("月",'/')
       w_ymd = w_ymd.replace("日",'')
       w_titleint = int(w_title[:4])
       w_title = str(w_titleint) + "年度決算"


       wb = op.load_workbook(export_file)
       sh_name = '矢崎総業'
       ws = wb[sh_name]
       ws.cell(row=max_row,column=2).value = w_title
       ws.cell(row=max_row,column=3).value = w_url
       ws.cell(row=max_row,column=4).value = w_ymd
       ws.cell(row=max_row,column=6).value = w_url
       ws.cell(row=max_row,column=6).hyperlink = w_url
       ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
       max_row += 1
       # エクセルファイルの保存
       wb.save(export_file)

