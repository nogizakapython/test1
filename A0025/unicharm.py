#######   ユニ・チャーム 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/04/08  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "unicharm" + date1 + ".txt"
out_file = "unicharm.txt"
date_str = ""
w_title = ""
base_url = 'https://www.unicharm.co.jp'
web_url = 'https://www.unicharm.co.jp/ja/ir/news/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
t_year = 0
year_array=[]

# Chromeを指定する
driver = webdriver.Chrome()

# エクセル出力関数 
def excel_output(w_title,w_url,w_ymd,max_row):
      wb = op.load_workbook(export_file)
      sh_name = 'ユニ・チャーム'
      ws = wb[sh_name]
      ws.cell(row=max_row,column=2).value = w_title
      ws.cell(row=max_row,column=3).value = w_url
      ws.cell(row=max_row,column=4).value = w_ymd
      ws.cell(row=max_row,column=6).value = w_url
      ws.cell(row=max_row,column=6).hyperlink = w_url
      ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
      # エクセルファイルの保存
      wb.save(export_file)


# Chromeを開いて企業HPにアクセスする
for year in [date5,date6]:
    target_url = web_url + str(year) + ".html"
   
    try:
        driver.get(target_url)
        sleep(3)

        for i in range(1,26):
            try:
                xpath_str1 = '/html/body/div[1]/main/div/div[2]/div/div/div/div/div/section/div[3]/div/div/div/ul/li[' + str(i) + ']'
                element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
                
            except:
                break    
        
          
       
    except EnvironmentError as e:
        str100 = e     
    except:
        str100 = 1
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
     

     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     if line1:
        row_count += 1
     else:
        break   

     result1 = re.search('a href',line1)
     result2 = re.match('                        <p>2...年..月..日',line1)
     result3 = re.match('                        <p>',line1)

     if result1:
        w_array1 = line1.split("=")
        w_urlstr = w_array1[1]
        w_urlstr = w_urlstr.replace('>','')
        w_urlstr = w_urlstr.replace('class','')
        w_urlstr = w_urlstr.replace('"','')
        w_url = base_url + w_urlstr
        # print(w_url)

     if result2:
        w_array2 = line1.split(">")
        w_ymdstr = w_array2[1]
        w_ymd = w_ymdstr.replace('</p','')
        # print(w_ymd)

     if result3:
        w_array3 = line1.split(">")
        w_titlestr = w_array3[1]
        w_titlestr = w_titlestr.replace('</p','')
        w_title = w_titlestr.replace('<sup','')
        # print(w_title)            
        key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
        title_result = re.search(key_word,w_title)
        if title_result:
           excel_output(w_title,w_url,w_ymd,max_row)
           max_row += 1


