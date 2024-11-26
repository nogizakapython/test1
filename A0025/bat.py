#######   BAT JAPAN 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/21  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys 
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')


input_file = "bat" + date1 + ".txt"
out_file = "bat.txt"
date_str = ""
w_title = ""
base_url = 'https://www.batj.com'
web_url = 'https://www.batj.com/media'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
html_array = ["#corp-tabs-9ee4e0f89c-item-ea1dd66584-tab","#corp-tabs-9ee4e0f89c-item-33996595c0-tab"]

# Chromeを指定する
driver = webdriver.Chrome()

i = 2
for html_name in html_array:
    # Chromeを開いて企業HPにアクセスする
    
    target_url = web_url + html_name + '.html'
    try:
        driver.get(target_url)
        sleep(5)
        for j in range(2,20):
            try:
                xpath_str1 = '/html/body/div[1]/div/div[4]/div/div/div[2]/div/div[' + str(i) + ']/div/div/ul/li[' + str(j) + ']/a'
                
            except:
                break    
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
            print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
        
        
    except EnvironmentError as e:
        str100 = e     
    except:
        str100 = ""
    i += 1


# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result1 = re.match('<a class',line1)
    result2 = re.match("          [^<]",line1)
#     result3 = re.match('             [^<]',line1)
    
    if result1:
       w_array1 = line1.split(">")
       w_ymdstr = w_array1[2]
       w_ymdstr = w_ymdstr.replace("</span","")
       w_ymdstr = w_ymdstr.replace('年',"/")
       w_ymdstr = w_ymdstr.replace('月',"/")
       w_ymdstr = w_ymdstr.replace('日',"")
       w_ymd = w_ymdstr.replace(" ","")
    #    print(w_ymd)
       w_array2 = line1.split("=")
       w_urlstr = w_array2[2]
       w_urlstr = w_urlstr.replace(" target","")
       w_urlstr = w_urlstr.replace('"','')
       w_url = base_url + w_urlstr
    #    print(w_url)

    if result2:
        w_titlestr = line1
        w_title = w_titlestr.replace(" ","")
        # print(w_title)
        key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート|経営)"
        title_result = re.search(key_word,w_title)
        if title_result:
            wb = op.load_workbook(export_file)
            sh_name = 'BATJAPAN'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
            ws.cell(row=max_row,column=6).hyperlink = w_url
            ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)

