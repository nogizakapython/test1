#######   EISAI 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/25  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')


input_file = "eisai" + date1 + ".txt"
out_file = "eisai.txt"
date_str = ""
w_title = ""
base_url = 'https://www.eisai.co.jp'
# web_url = 'https://www.eisai.co.jp/ir/library/settlement/index.html'
web_url = 'https://www.eisai.co.jp/ir/library/index.html'


max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
w_titlehead = ""
flag1 = 0


# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
    
target_url = web_url
driver.get(target_url)

try:
    driver.get(target_url)
    sleep(10)
    
    try:
        xpath_str0 = '/html/body/div[1]/div[3]/div/main/h3'
        element_str0 = driver.find_element(by=By.XPATH,value=xpath_str0)
        print(element_str0.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
    except:
        print("タグ取得失敗")

    for i in range(1,4):
        try:
            # xpath_str1 = '//*[@id="tgl-sp-' + str(i) + '"]/div[2]/table/tbody/tr[' + str(j) + ']' 
            xpath_str1 = '/html/body/div[1]/div[3]/div/main/ul[1]/li[' + str(i) + ']'
                         
        except:
            break    
        element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
      
       
except EnvironmentError as e:
    str100 = e     
except:
    str100 = ""
     

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
      row_count += 1
    else:
      break   

    # 2025/8/6 検索条件変更に伴う修正 takao.hattori
    result1 = re.match('<h3',line1)
    result2 = re.match('<p>2',line1)
    result3 = re.match('<p><a class',line1)
    result4 = re.match('<a href',line1)
   
    if result1:
        w_array0 = line1.split('>')
        w_titlehead = w_array0[1]
        w_titlehead = w_titlehead.replace('</h3','')

    if result2:
       w_array1 = line1.split(">")
       w_ymd = w_array1[1]
       w_ymd = w_ymd.replace('</p','')
       w_ymd = w_ymd.replace('年','/')
       w_ymd = w_ymd.replace('月','/')
       w_ymd = w_ymd.replace('日','')
       ymd_array = w_ymd.split("/")
       year1 = ymd_array[0]
       month1 = int(ymd_array[1])
       day1 = int(ymd_array[2])
       if month1 < 10:
           month1 = "0" + str(month1)
       else:
           month1 = str(month1)
       if day1 < 10:
           day1 = "0" + str(day1)
       else:
           day1 = str(day1)
       w_ymd = year1 + "/" + month1 + "/" + day1            
           
    #    print(w_ymd)
    # URLとタイトルが、タグで2種類存在するため、修正(2025/8/6 takao.hattori) 
    if result3:
        w_array4 = line1.split('>')
        w_titlestr = w_array4[2]
        w_titlestr = w_titlestr.replace('<span class="icon-pdf"','')
        w_title = w_titlehead + " " + w_titlestr
        # print(w_title)

        w_array5 = line1.split("=")
        w_urlstr = w_array5[2]
        w_urlstr = w_urlstr.replace(' target','')
        w_urlstr = w_urlstr.replace(' class','')
        w_urlstr = w_urlstr.replace('"','')
        w_url = base_url + w_urlstr
        # print(w_url)

        flag1 = 1


    if result4:
        w_array2 = line1.split('>')
        w_titlestr = w_array2[1]
        w_titlestr = w_titlestr.replace('<span class="icon-pdf"','')
        w_title = w_titlehead + " " + w_titlestr
        # print(w_title)

        w_array3 = line1.split("=")
        w_urlstr = w_array3[1]
        w_urlstr = w_urlstr.replace(' target','')
        w_urlstr = w_urlstr.replace(' class','')
        w_urlstr = w_urlstr.replace('"','')
        w_url = base_url + w_urlstr
        # print(w_url)
        flag1 = 1
    
    # 2025/8/6 Excelにレコードを1件出力する条件を追加 takao.hattori
    if flag1 == 1:
        key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート)"
        title_result = re.search(key_word,w_title)
        if title_result:
           wb = op.load_workbook(export_file)
           sh_name = 'EISAI'
           ws = wb[sh_name]
           ws.cell(row=max_row,column=2).value = w_title
           ws.cell(row=max_row,column=3).value = w_url
           ws.cell(row=max_row,column=4).value = w_ymd
           ws.cell(row=max_row,column=6).value = w_url
           ws.cell(row=max_row,column=6).hyperlink = w_url
           ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
           max_row += 1
           flag1 = 0
           # エクセルファイルの保存
           wb.save(export_file)

