#######   ノバルティスファーマ 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/29  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font
from selenium.webdriver.support.ui import Select


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
youbi = dt.strftime('%a')
date5 = int(date2)
date6 = date5 - 1
date7 = dt.strftime('%Y/%m/%d')

input_file = "novartis" + date1 + ".txt"
out_file = "novartis.txt"
date_str = ""
w_title = ""
base_url = 'https://www.novartis.com'
web_url = 'https://www.novartis.com/investors/financial-data/annual-results'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
t_year = 0
year_array=[]

# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする

def output_excel_file(max_row,w_title,w_ymd,w_url):
    wb = op.load_workbook(export_file)
    sh_name = 'ノバルティス'
    ws = wb[sh_name]
    ws.cell(row=max_row,column=2).value = w_title
    ws.cell(row=max_row,column=3).value = w_url
    ws.cell(row=max_row,column=4).value = w_ymd
    ws.cell(row=max_row,column=6).value = w_url
    ws.cell(row=max_row,column=6).hyperlink = w_url
    ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
    wb.save(export_file)


target_url = web_url
try:
   driver.get(target_url)
   sleep(3)

   try:
      xpath_str1 = '//*[@id="tabannual-results-16991"]/div/div/article/div'
   except:
      print("データ取得に失敗しました",file=codecs.open(input_file,'a','utf-8'))
      
   element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
   print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
       
except EnvironmentError as e:
   str100 = e     
except:
   str100 = 1
     
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    w_urlstr = ""
    w_titlestr = ""
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
      row_count += 1
    else:
      break   

    result1 = re.match('<div class="clearfix',line1)
    

    if result1:
       w_array1 = line1.split(">")
       w_titlestr1 = w_array1[2]
       w_titlestr1 = w_titlestr1.replace('</h2','')
       ymd_array1 = w_titlestr1.split('-')
       
       w_titlestr1 = ymd_array1[0]
       w_ymd1 = ymd_array1[1]
       w_ymd1 = w_ymd1.replace('&nbsp;','/')
       w_ymd1 = w_ymd1.replace(',','')
      
       w_titlestr2 = w_array1[4]
       w_titlestr2 = w_titlestr2.replace('</h3','')
       w_urlstr1 = w_array1[6]
       w_urlstr1 = w_urlstr1.replace(' target="_blank"','')
       w_urlstr1 = w_urlstr1.replace('<a href=','')
       w_urlstr1 = w_urlstr1.replace('"','')
       w_title1 = w_titlestr1 + " " + w_titlestr2
       w_url1 = base_url + w_urlstr1
       output_excel_file(max_row,w_title1,w_ymd1,w_url1)
       max_row += 1

       w_urlstr2 = w_array1[14]
       w_urlstr2 = w_urlstr2.replace('<a href=','')
       w_urlstr2 = w_urlstr2.replace('target="_blank"','')
       w_urlstr2 = w_urlstr2.replace('"','')
       w_url2 = base_url + w_urlstr2
      
       w_titlestr3 = w_array1[15]
       w_titlestr3 = w_titlestr3.replace('</a','')
       w_titlestr3 = w_titlestr3.replace('&nbsp','')
       w_titlestr3 = w_titlestr3.replace(';',' ')
       w_titlestr3 = w_titlestr3.replace('"','')
       w_title2 = w_titlestr1 + " " + w_titlestr3
       
       output_excel_file(max_row,w_title2,w_ymd1,w_url2)
       max_row += 1
       

