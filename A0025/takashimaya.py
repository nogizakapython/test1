######   高島屋 中計・決算ニュース情報取得ツール　###########
#######   新規作成  2024/04/03  ##########
#######   Author  takao.hattori ###########

#ライブラリのインポート
from bs4 import BeautifulSoup
import urllib3
import codecs
import datetime  
import re
import os
import requests
import shutil
from openpyxl.styles.fonts import Font
import openpyxl as op

#IRニュース取得先URL変数の定義
b_url = "https://www.takashimaya.co.jp/corp/topics/index.html?category=IR"
# ベースURL
base_url = "https://www.takashimaya.co.jp"
#開始ページ変数の定義
start_num = 1
#終了ページ変数の定義
end_num = 2
# 現在日時の取得
dt = datetime.datetime.now()
#　現在日時を年4桁、月2桁、日付2桁、時間、分、秒のフォーマットで取得する
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
# 格納先ファイル名の定義
file_name = "takashimaya" + date1 + ".txt"

#検索文字の設定(日付タグ)
patturn1 = '<dt class="date">'


#URL、ニュースリリース時間の設定
patturn2 = '<dd class="txt">'

input_file = "takashimaya" + date1 + ".txt"
output_file = "takashimaya.txt"

http = urllib3.PoolManager()
url = b_url
row_count = 0
row_count2 = 0
start_str = '					<li>'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"


def remove_output_file():
   os.remove(output_file)

def output_work_file(data_record):
   print(data_record,file=codecs.open(output_file,'a','utf-8'))
   

# スクレイピング対象の URL にリクエストを送り HTML を取得する
r = requests.get(url)
soup = BeautifulSoup(r.text, 'html.parser')
print(r.text,file=codecs.open(input_file,'a','utf-8'))
fileobj1 = open(input_file,encoding='utf-8')
start_flag = 0

output_file_exist = os.path.isfile(output_file)
if output_file_exist:
   remove_output_file()


# 全体から最新のニュース
while True:
   line1 = fileobj1.readline()
   line1 = line1.replace("\n","")
   #取得したHTMLから、必要なデータを抽出し、抽出ファイルに書き込む
   if line1 == start_str:
      output_work_file(line1)
      start_flag = 1
   elif start_flag == 1: 
      output_work_file(line1)
      row_count += 1
   
   if row_count > 1000:
      break
fileobj1.close()

fileobj2 = open(output_file,encoding="utf-8")
while True:
     w_urlstr = ""
     w_titlestr = ""
     line2 = fileobj2.readline()
     if line2:
        row_count2 += 1
     else:
        break   
     
     result1 = re.search(patturn1,line2)
     result2 = re.search(patturn2,line2)

     if result1:
        w_array1 = line2.split(">")
        w_ymdstr = w_array1[1]
        w_ymdstr = w_ymdstr.replace('</dt','')
        w_ymd = w_ymdstr.replace('.','/')
      #   print(w_ymd)

     if result2:
        w_array2 = line2.split(">")   
        w_urlstr = w_array2[1]
        url_array = w_urlstr.split('=')
        w_urlstr = url_array[1]
        w_urlstr = w_urlstr.replace('"','')
        w_url = base_url + w_urlstr
      #   print(w_url)

        w_titlestr = w_array2[2]
        w_title = w_titlestr.replace('</a','')
      #   print(w_title)

        key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|Financial Results)"
        title_result = re.search(key_word,w_title)
        if title_result:
            wb = op.load_workbook(export_file)
            sh_name = '高島屋'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
            ws.cell(row=max_row,column=6).hyperlink = w_url
            ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)
