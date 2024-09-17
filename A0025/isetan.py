#######   三越・伊勢丹HD 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/26  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
from openpyxl.styles.fonts import Font
import selenium

dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
date4 = dt.strftime('%Y年%m年%d日')
date5 = int(date2)

input_file = "isetan" + date1 + ".txt"
out_file = "isetan.txt"
date_str = ""
w_title = ""
base_url = 'https://www.imhds.co.jp'
web_url = 'https://www.imhds.co.jp/ja/ir/ir_news/archive.html?year='
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
w_url = ""
w_titlehead = ""


# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
for year in  [date5]:
   target_url = web_url + str(year) + '.html'
   try:
      driver.get(target_url)
      sleep(3)

      for i in range(1,31):
         try:
            
            xpath_str1 = '/html/body/div[1]/div/div[3]/div/div[1]/div/div/div[2]/section[1]/ul/li[' + str(i) + ']'
            

         except:
            break    
         element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
         print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))
   
   except selenium.common.exceptions.NoSuchElementException as e:
      str100 = e
   except:
      str100 = "OK"   

     

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
   line1 = fileobj.readline()
   line1 = line1.replace("\n","")
   if line1:
      row_count += 1
   else:
      break   
   url_result = re.match('			<a href',line1)
   if url_result:
      w_array1  = line1.split("=")
      w_urlstr = w_array1[1]
      w_urlstr = w_urlstr.replace('target',"")
      w_urlstr = w_urlstr.replace('"display: block;" data-year','')
      w_urlstr = w_urlstr.replace('"','')
      https_result = re.search('https',w_urlstr)
      if https_result:
         w_url = w_urlstr
      else:
         w_url = "https:" + w_urlstr   
      # print(w_url) 

      w_array2 = line1.split(">")
      w_ymd = w_array2[3]
      w_ymd = w_ymd.replace('</time','')
      w_ymd = w_ymd.replace('年','/')
      w_ymd = w_ymd.replace('月','/')
      w_ymd = w_ymd.replace('日','')
      # print(w_ymd)
       
       
   time_urlresult = re.match('			<div class="c-list-post__wrap">',line1)
   if time_urlresult:
      w_array3 = line1.split(">")
      w_ymd = w_array3[2]
      w_ymd = w_ymd.replace('</time','')
      w_ymd = w_ymd.replace('年','/')
      w_ymd = w_ymd.replace('月','/')
      w_ymd = w_ymd.replace('日','')
      w_url = ""
      # print(w_ymd)
   
   title_result1 = re.match('				<p class="c-list-post__text">',line1)
   if title_result1:
      w_array4 = line1.split(">")
      w_titlestr = w_array4[1]
      w_titlestr = w_titlestr.replace('<span','')
      w_titlestr = w_titlestr.replace('</p','')
      w_title = w_titlestr.replace('<br','')
      # print(w_title)
   

      key_word = r"(決算|株主総会|説明会|IR説明会|中期経営計画|報告書|レポート|経営)"
      title_result = re.search(key_word,w_title)
      if title_result:
         wb = op.load_workbook(export_file)
         sh_name = '三越伊勢丹ホールディングス'
         ws = wb[sh_name]
         ws.cell(row=max_row,column=2).value = w_title
         ws.cell(row=max_row,column=3).value = w_url
         ws.cell(row=max_row,column=4).value = w_ymd
         ws.cell(row=max_row,column=6).value = w_url
         ws.cell(row=max_row,column=6).hyperlink = w_url
         ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')
                
         max_row += 1
         # エクセルファイルの保存
         wb.save(export_file)
