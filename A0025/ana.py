#######   ANA 決算ニュース情報取得ツール　###########
#######   新規作成  2024/03/19  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from openpyxl.styles.fonts import Font


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')
w_ymd = dt.strftime('%Y年%m年%d日')


input_file = "ana" + date1 + ".txt"
out_file = "ana.txt"
date_str = ""
w_title = ""
base_url = 'https://www.ana.co.jp'

web_url = 'https://www.ana.co.jp/group/investors/irdata/disclosure/'
max_row = 5
base_file = "【IR】検索結果_yyyymmdd.xlsx"
export_file = "【IR】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""


# Chromeを指定する



driver = webdriver.Chrome()


target_url = web_url
try:
    driver.get(target_url)
    sleep(5)
    # 2026/8/25 takao.hattori xpath変更に伴う修正
    for i in range(1,3):
        for j in range(1,31):
            k = 2 * i - 1
            xpath_str1 = '//*[@id="mainContent"]/div[2]/div/div/div/div[' + str(k) + ']/div[1]/section[1]/div[' + str(i) + ']/ul/li[' + str(j) + ']'
            
            try:
                element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
            except:
                str1 = "data finish"
                break
            print(element_str1.get_attribute("outerHTML"),file=codecs.open(input_file,'a','utf-8'))

except EnvironmentError as e:
    str100 = e     
except:
    str1000 = ""

# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(input_file,out_file)


fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    # 2026/8/25 TAKAO.HATTORI 検索条件変更に伴う修正
    result1 = re.search('s_eirModule_date_time',line1)
    result2 = re.search('<a',line1)
    result3 = re.search('s_titleBox_title_link_label',line1)
  
    # 2026/8/25 ニュース掲載日付の条件変更の伴う修正
    if result1:
       w_array1 = line1.split(">")
       w_ymd = w_array1[1]
       w_ymd = w_ymd.replace('</time',"")
       w_ymd = w_ymd.replace('.',"/")
    #    print(w_ymd)

    # 2026/8/25 URLの条件変更の伴う修正
    if result2:   
       w_array2 = line1.split("=")
       w_url = w_array2[2]
       w_url = w_url.replace(' target','')
       w_url = w_url.replace('"','')
    #    print(w_url)

    # 2026/8/25 ニュースタイトルの条件変更の伴う修正    
    if result3:
       w_array3 = line1.split(">")
       w_title = w_array3[1]
       w_title = w_title.replace('</span',"")
    #    print(w_title)
       key_word = r"(決算|株主総会|説明会|IR説明会|中期経営|報告書|レポート|経営)"
       title_result = re.search(key_word,w_title)
       if title_result:
            wb = op.load_workbook(export_file)
            sh_name = 'ANA'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
            ws.cell(row=max_row,column=6).hyperlink = w_url
            ws.cell(row=max_row,column=6).font = Font(color='0000FF',underline='single')      
            max_row += 1
            wb.save(export_file)
            
