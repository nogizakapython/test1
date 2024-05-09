############ 中計決算50社 管理用エクセルファイルにニュースリリース日 ###########
############ 対象データを書き込む処理 ######################################
############ 新規作成  2024/4/24      ######################################
############ 作成者    takao.hattori  ######################################
############################################################################

# ライブラリのインポート
import openpyxl as op
from datetime import datetime
from openpyxl.styles.fonts import Font
from datetime import timedelta
import shutil
# 日付の定義
dt = datetime.now()
date1 = dt.strftime('%Y%m%d')
date2 = dt.strftime('%Y/%m/%d')
# 更新ファイルの定義
base_file = "A0055_中計決算50社_テンプレートファイル.xlsx"
update_file = "A0055_中計決算50社_" + str(date1) + ".xlsx"
shutil.copy(base_file,update_file)
wb1 = op.load_workbook(update_file)
sh_name = '更新データ'
ws1 = wb1[sh_name]
w1_row_count = 6
# 変数の定義
w_title = ""
w_url = ""
# 書き込み元のファイルの変数定義
input_file = "【IR】検索結果_" + date1 + ".xlsx"
wb2 = op.load_workbook(input_file)
sheet_array1 = ["AEON","味の素","ANA","ASAHI","アステラス製薬","BRIDGESTONE","BRISTOL-MYERS"
                ,"BATJAPAN","中外製薬","日本コカ・コーラ","マツキヨ","三菱ふそう","大東建託",
                "DAIWA","DENSO","EISAI","RETAILING","ホンダ","三越伊勢丹ホールディングス",
                "日本航空","JTB","KAO","KUBOTA","LIXIL","三菱商事","日本ハム","日本通運","NITORI",
                "日本郵船","小田急電鉄","オリエンタルランド","オムロン","日産自動車","良品計画",
                "参天製薬","セブン＆アイ・ホールディングス","シマノ","塩野義製薬","資生堂","SUBARU",
                "大正製薬","高島屋","武田薬品工業","トヨタ自動車","トヨタ紡織","ユニ・チャーム",
                "矢崎総業"]

# 前回の作業日を入力する関数
def data_input():
    #　前回の作業入力ルーチン
    
    ymd = ""
    while True:
        corect_count = 0
        print("前回の作業日付を西暦4桁、月2桁、日付2桁で入力してください。例) 2024年5月8日の場合は「20240508」と入力してください\n")
        data1 = input()
        if len(data1) == 8:
            corect_count += 1
        else:
            print("入力文字数が違います。")
        
        w_year = data1[:4]
        w_month = data1[4:6]
        w_day = data1[6:]
        if int(w_month) < 1 or int(w_month) > 12:
            print("NG")
        else:     
            corect_count += 1
    
        if int(w_day) < 1 or int(w_day) > 31:
            print("NG")
        else:     
            corect_count += 1
        if corect_count == 3:
            ymd = w_year + '/' + w_month + '/' + w_day
            break 
    return ymd        


  
# 前回納品日に転記した最終日と今回納品分の転記最終日を更新エクセルファイルに出力する関数
def finish_write_day(before_ymd,now_ymd):
    ws1.cell(row=2,column=4).value = before_ymd
    ws1.cell(row=3,column=4).value = now_ymd
    wb1.save(update_file)  


# アップデートエクセルファイルの書き込み開始行を取得する関数
def update_excel_maxcount():
    update_row_count = 6
    while True:
        value1 = ws1.cell(row=update_row_count,column=3).value
        if value1 == None:
            break
        else:
            update_row_count += 1
            # print(admin_row_count)
    return update_row_count    

# 企業別前回のニュースリリース日を先週分の最終リリース日のセルに書き込む関数


# メイン処理
str_ymd = data_input()


# 今回の作業日と前回の作業日の日付差を取得し、前回の作業日から今回の作業日の前日まで日付を
# 配列に格納する。
# print(str_ymd)
ymd = datetime.strptime(str_ymd, '%Y/%m/%d')
finish_write_day(ymd,date2)
array1 = []
dt1 = datetime.now()
day_sabun = dt1 - ymd
day_sabun = str(day_sabun)
sabun_array1 = day_sabun.split(" ")
day_sa = sabun_array1[0]
day_sa = int(day_sa)

# 前回の作業日翌日から今回の作業日までを日付配列に格納する。
for i in range(0,day_sa):
    w_ymd = ymd + timedelta(i)
    w_ymd = w_ymd.strftime("%Y/%m/%d")
    array1.append(w_ymd)
    

# 管理用エクセルファイルの書き込み開始行を取得する
w1_row_count = update_excel_maxcount()

# 中計・決算自動化対象48社別シートからスクレイピングデータを読み込む 
for sh_name in list(sheet_array1):
    ws2 = wb2[sh_name]
    w2_row_count = 5
    
    # 各企業別のエクセルシートの最終行を取得する
    while True:
        data_value = ws2.cell(row=w2_row_count,column=2).value
        if data_value == None:
            break
        else:
            w2_row_count += 1
    #管理用シートに書き込むデータをニュースリリース日と前回の作業日の翌日から今日までの日付で
    #突き合わせして、合致したものを管理用エクセルファイルに書き込む  
    
    for row_count in range(w2_row_count-1,4,-1):
        ad_company = sh_name
        ad_title = ws2.cell(row=row_count,column=2).value
        ad_url = ws2.cell(row=row_count,column=3).value
        ad_ymd = ws2.cell(row=row_count,column=4).value
        
        for t_ymd in array1:
            if ad_ymd == t_ymd:
                ws1.cell(row=w1_row_count,column=3).value = ad_company
                ws1.cell(row=w1_row_count,column=4).value = ad_ymd
                ws1.cell(row=w1_row_count,column=5).value = ad_title
                ws1.cell(row=w1_row_count,column=5).hyperlink = ad_url
                ws1.cell(row=w1_row_count,column=5).font = Font(color='0000FF',underline='single',name='Meiryo UI')
                w1_row_count += 1
                wb1.save(update_file)    
    
    