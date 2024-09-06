############ myTE承認、SVレビュー承認済リスト作成処理 ###########
############ 新規作成  2024/8/8      ######################################
############ 修正  2024/8/23  名前のスペース対応      ###################
############ 修正  2024/9/6  稲葉さんを追加      ###################
############ 作成者    takao.hattori  ######################################
############################################################################

# ライブラリの読み込み
import re
import shutil
import openpyxl as op
from datetime import datetime
import sys

# 入力ファイル(Teamsのスレッドから手動でテキストファイルをコピー)
input_file = "myTE.txt"
# テンプレートエクセルファイル
template_file = "myTE_SVレビュー完了者一覧テンプレートファイル.xlsx"
dt = datetime.now()
date1 = dt.strftime('%Y%m%d')


#　入力ファイルの行数チェック関数
def inputfile_count(file_name):
    with open(file_name,encoding='utf-8') as f:
        count = 0
        for line in f:
            count += 1
        return count    

# XチームメンバーEID最終行取得関数
def member_count(output_file,start_num_count):
    # 出力ファイルの定義
    wb = op.load_workbook(output_file)
    sh_name = "チェックシート"
    ws = wb[sh_name]

    # 出力ファイル開始行
    i = start_num_count
    while True:
        eid_check = ws.cell(row=i,column=1).value
        if eid_check == None:
            break
        else:
            i += 1
    return i

# 〇印を対象EIDの方の3列目に書き込む関数
def output_status(output_file,start_num,end_num,eid):
    # 出力ファイルの定義
    wb = op.load_workbook(output_file)
    sh_name = "チェックシート"
    ws = wb[sh_name]
    for i in range(start_num,end_num+1):
        w_eid = ws.cell(row=i,column=1).value
        if eid == w_eid:
            ws.cell(row=i,column=3).value = "〇"
    wb.save(output_file)    
          
         
def main():
    # 入力ファイル(Teamsのスレッドから手動でテキストファイルをコピー)
    input_file = "myTE.txt"
    # テンプレートエクセルファイル
    template_file = "myTE_SVレビュー完了者一覧テンプレートファイル.xlsx"
    dt = datetime.now()
    date1 = dt.strftime('%Y%m%d')
    # myTE自動化ツールコピー先ファイル
    output_file = "myTE_SVレビュー完了者一覧" + date1 + ".xlsx"
    


    # Teamsで表示される名前とEIDを紐づける連想配列(メンバーが変わるたびに修正)
    # 2024/9/6 稲葉さんを追加
    xteam_member = {
                "Wakabayashi, Eri" : "eri.wakabayashi",
                "Kimura, Yuichi" : "yuichi.kimura",
                "Watanabe, Rintaro" : "rintaro.watanabe",
                "Kurematsu, Takumi" : "takumi.kurematsu",
                "Maeda, Shingo" : "shingo.a.maeda",
                "Tanikawa, Hironori" : "hironori.tanikawa",
                "Anzai, Keita" : "keita.anzai",
                "Motoyama, Kazuma" : "kazuma.motoyama",
                "Hattori, Takao" : "takao.hattori",
                "Hasegawa, Rei" : "rei.hasegawa" ,
                "Terui, Yumi" : "yumi.terui" ,
                "Taniguchi, Shunsuke" : "shunsuke.taniguchi" ,
                "Nakano, Tomoki" : "tomoki.nakano" ,
                "Nozawa, Ryota" : "ryota.nozawa" ,
                "Daikusu, Riamu" : "riamu.daikusu" ,
                "Ohura, Kentaro" : "kentaro.ohura" ,
                "Fukazawa, Takuma" : "takuma.fukazawa" ,
                "Muramatsu, Takashi" : "takashi.muramatsu" ,
                "Nishikawa, Kazuma" : "kazuma.nishikawa" ,
                "Eto, Shinsuke" : "shinsuke.eto" ,
                "Anzai, Keita" : "keita.anzai" ,
                "Inaba, Ryusei" : "ryusei.inaba"
                }

    # 出力ファイル開始行
    start_num = 4
    # 出力ファイル終了行(メンバー変更のたびに変更)
    input_file_row = 0
    # Excelファイルをコピー
    try:
        shutil.copy(template_file,output_file)
    except PermissionError:
        print(f"{template_file}または{output_file}が開いています。ファイルを閉じて、再実行してください。")
        sys.exit()

    end_num = member_count(output_file,start_num)

    fileobj = open(input_file,encoding="utf-8")

    input_file_row = inputfile_count(input_file)
    
    file_row_count = 0
    while True:
         line1 = fileobj.readline()
         line1 = line1.replace("\n","")
        
         if line1 == "" and file_row_count == input_file_row :
            break
        
         result1 = re.search("SVレビュー完了しました。",line1)
         if result1:
             array1 = line1.split(" ")
             name1 = array1[0] + " " + array1[1]
             result2 = re.search("SVレビュー完了しました。",line1)
             if result2:
                 name1 = name1.replace("SVレビュー完了しました。Submitを押してください。","")
                 name1 = name1.replace("SVレビュー完了しました。Submit押してください。","")
                 name1 = name1.replace('　','')
                 eid = xteam_member[name1]
                 
                 output_status(output_file,start_num,end_num,eid)           
                            
         file_row_count += 1
         

if __name__ == "__main__":
    main()        