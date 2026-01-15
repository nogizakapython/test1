#######   ICEnergy 石油平均価格推移情報取得ツール　###########
#######   新規作成  2026/01/14  ##########
#######   Author  takao.hattori ###########


# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
# opempyxlライブラリをインポート
import openpyxl as op
# 待機時間ライブラリをインポート
from time import sleep


# 取得先のURLの変数
target_url = 'https://pps-net.org/statistics/crude-oil'
# 出力先のExcelファイル変数(マクロ作成時にファイル名を変更)
export_file = "石油価格値段.xlsx"


# xpath取得情報配列の定義
xpath_array = []

# 取得結果格納配列
data_array = []

# 出力先のエクセルファイルの変数定義
wb = op.load_workbook(export_file)
sh_name = 'Sheet1'
ws = wb[sh_name]


# Excel出力関数
def output_file(output_data,row_num,col_num):
    if row_num == 1:
        ws.cell(row=row_num,column=col_num).value = output_data
    else:
        ws.cell(row=row_num,column=col_num).value = float(output_data)    
    wb.save(export_file)
    
# 作業対象月のデータクレンジング処理
def work_month_data_cleansing(ymd):
    # 余分な文字列を削除
    ymd = ymd.replace("<th>","")
    ymd = ymd.replace("</th>","")
    w_array = ymd.split('/')
    month = w_array[1]
    work_month = ""
    # 作業対象月が1月の時は西暦を付けてYYYY年MM月に、1月以外はMM月に変換する
    if month == "1":
        work_year = w_array[0]
        work_month = work_year + "年" + month + "月"
    else:
        work_month = month + "月"
    # 文字列を返す
    return work_month

# 価格データのデータクレンジング処理
def data_clansing(work_data_accout):
    work_data_accout = work_data_accout.replace(' $/バレル',"")
    work_data_accout = work_data_accout.replace('<td>',"")
    work_data_accout = work_data_accout.replace('</td>',"")
    # 不要な文字を取り除いたデータを返す
    return work_data_accout

# メイン関数
def main():
    # 出力開始行の変数
    start_row_num = 1
    # 出力先の列数変数
    col_num = 2
    # 出力先の列数変数
    col_num = 2
    row_num = start_row_num
    # Chromeを指定する
    driver = webdriver.Chrome()

    try:
        driver.get(target_url)
        sleep(1)

        try:
            xpath_str1 = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[1]/th[2]'
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)

            for i in range(2,6):
                xpath_str = '/html/body/div[4]/div[1]/div[1]/div/div/table[1]/tbody/tr[' + str(i) + ']/td[1]'
                element_str = driver.find_element(by=By.XPATH,value=xpath_str)
                data1 = element_str.get_attribute("outerHTML")
                data_array.append(data1)
                   
        except:
            print("配列エラー")
        
    except EnvironmentError as e:
        print("作業対象月の原油価格を取得できませんでした")
        exit()     
    
    
    ymd = element_str1.get_attribute("outerHTML")
    work_month = work_month_data_cleansing(ymd)
    output_file(work_month,row_num,col_num)
    row_num += 1

    for j in range(0,4):
        data_accout = data_array[j]
        data_accout = data_clansing(data_accout)
        output_file(data_accout,row_num,col_num)
        row_num += 1
    

if __name__ == "__main__":
    main()
    