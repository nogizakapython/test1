####################################################
####       原因別FB件数取得スクリプト          ######
####    新規作成  :  2024/4/10                ######
####    Create By  takao.hattori              ######
####################################################
# ライブラリ
import os 
import glob
import codecs
import openpyxl as op
import re
import datetime 


output_file = "filelist.txt"
fb_reason_count = 3
fb_data_count = 0
fb_array = []
fb_count = 0
fb_data_count = 0
start_fb_count = fb_data_count
end_fb_count = 0
project_name = ""
reason_count = 0
insident_count = 0
process_str = ""
result_file = ""


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')


file_exist = os.path.isfile(output_file)
if file_exist:
    os.remove(output_file)

# 案件リスト
project_array = ["A0048","SN_0025","SN_0031","SN_0052"]
for file in project_array:
    file_patturn1 = file + "*.xlsx"
    test1 = glob.glob(file_patturn1)
    for str1 in test1:
        print(str1,file=codecs.open(output_file,'a','utf-8'))

with open(output_file,encoding="utf-8",mode="r") as f:
    feedback_sheetname = 'FB'
    list_sheetname = 'List'
    while True:
        file_name = f.readline()
        file_name = file_name.replace('\n',"")
        A0048_flag = re.search("A0048",file_name)
        SN_0025_flag = re.search("SN_0025",file_name)
        SN_0031_flag = re.search("SN_0031",file_name)
        SN_0052_flag = re.search("SN_0052",file_name)
        if A0048_flag:
            project_name = "空調設備サーバールーム"
        if SN_0025_flag:
            project_name = "人事異動"
        if SN_0031_flag:
            project_name = "I&D"
        if SN_0052_flag:
            project_name = "JPiT myTE"            
        if file_name == "":
           break
        
        wb = op.load_workbook(file_name)
        wl = wb[list_sheetname]
        wf = wb[feedback_sheetname]
            
        while True:
            if  project_name == "JPiT myTE":
                reason_str = wl.cell(row=fb_reason_count,column=12).value
                fb_data_count = 10
            elif  project_name == "I&D":
                reason_str = wl.cell(row=fb_reason_count,column=11).value
                fb_data_count = 10    
            else:
                reason_str = wl.cell(row=fb_reason_count,column=11).value
                fb_data_count = 9
            start_fb_count = fb_data_count        
            if reason_str == None:
                break
            else:
                fb_array.append(reason_str)
                fb_reason_count += 1

        while True:
            if project_name == "I&D" or project_name == "JPiT myTE":
                work_name = wf.cell(row=5,column=5).value
                process_str = wf.cell(row=4,column=5).value
            else:
                work_name = wf.cell(row=4,column=5).value
                 
            output_file =  project_name + "名前" + work_name + date1 + ".csv"  
            data_str = wf.cell(row=fb_data_count,column=11).value
            if data_str == None:
                end_fb_count = fb_data_count
                break
            else:
                fb_data_count += 1
        for reason in fb_array:
            for i in range(start_fb_count,end_fb_count+1):
                data_str = wf.cell(row=i,column=11).value
                if data_str == reason:
                    reason_count += 1
            if project_name == "I&D" or project_name == "JPiT myTE":
                msg = project_name + "," + process_str + "," + work_name + "," + reason + "," + str(reason_count)
            else:                
                msg = project_name + "," + work_name + "," + reason + "," + str(reason_count)
            print(msg,file=codecs.open(output_file,'a','shift-jis')) 
            reason_count = 0

        for i in range(start_fb_count,end_fb_count+1):
            reason = "インシデント"
            insident_str = wf.cell(row=i,column=12).value
            if insident_str == "〇":
                insident_count += 1
        if project_name == "I&D" or project_name == "JPiT myTE":
            msg = project_name + "," + process_str + "," + work_name + "," + reason + "," + str(insident_count)
        else:                
            msg = project_name + "," + work_name + "," + reason + "," + str(insident_count)
        print(msg,file=codecs.open(output_file,'a','shift-jis')) 
        insident_count = 0         



