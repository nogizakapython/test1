#######   ICEnergy 石油平均価格推移情報取得ツール(共通クラス)　###########
#######   新規作成  2026/01/15  ##########
#######   Author  takao.hattori ###########


# opempyxlライブラリをインポート
import openpyxl as op
# 出力先ファイル
export_file = "月別石油価格推移データワークシート.xlsx"
# 出力先のエクセルファイルの変数定義
try:
    wb = op.load_workbook(export_file)
    sh_name = 'Sheet1'
    ws = wb[sh_name]
except:
    print(f"{export_file}が開いているか、フォルダ内に{export_file}がありませんのどちらかです。ファイルが開いているか、同じフォルダ内の存在しているか確認し、再実行してください") 
    exit(1)

# 共通クラス
class Common_data:
    # データ削除メソッド
    def data_clear(self):
        start_row = 1
        end_row = 5
        start_col = 1
        end_col = 25
        for i in range(start_row,end_row + 1):
            for j in range(start_col,end_col + 1):
                ws.cell(row=i,column=j).value = ""
        wb.save(export_file)
        wb.close()
    # Excelにデータを出力するメソッド
    def output_data(self,output_data,row_num,col_num):
        if row_num == 1:
            ws.cell(row=row_num,column=col_num).value = output_data
        else:
            ws.cell(row=row_num,column=col_num).value = float(output_data)    
        wb.save(export_file)
        wb.close()
    