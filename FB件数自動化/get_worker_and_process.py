####################################################
####    作業者、工程を取得するクラス         　######
####    新規作成  :  2024/4/12                ######
####    Create By  takao.hattori              ######
####################################################
####    クラスの使い方
####    ①インスタンス変数にFBシートのファイル名を引数に指定して呼び出す
####    ex
####    get_process_worker = Get_worker_and_process(file_name)
####    ②get_process_workerメソッドを呼ぶ
####    get_process_worker.get_process_worker()
####
####
####################################################


# Get_workerクラスを呼び出し、クラスを継承する。
from get_worker import Get_worker

class Get_worker_and_process(Get_worker):

    def get_worker(self):
        import openpyxl as op
        feedback_sheetname = 'FB'
        wb = op.load_workbook(self.__file_name__)
        wf = wb[feedback_sheetname]
        process_name = wf.cell(row=4,column=5).value
        work_name = wf.cell(row=5,column=5).value
        return_data = work_name + "," + process_name
        return_data_array = return_data.split(",")
        return return_data_array    