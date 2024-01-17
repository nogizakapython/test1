###################################################################
###   日経サイト検索対象、エクセルファイルへの書き込み処理 ##########
###   新規作成 2023/12/13 takao.hattori  ##########################
###   修正     2024/1/17  takao.hattori  URL活性化処理を追加   #####
###################################################################

#ライブラリの読み込み
import openpyxl as op
import datetime
import sys
import shutil


#今日の日付を取得
date1 = datetime.datetime.today()
#今日の日付を年、月、日で分ける
t_year = str(date1.year)
t_month = str(date1.month)
t_day = str(date1.day)
#月、日が1桁の時、0埋めをして2桁にする。 
t_month = t_month.rjust(2,'0')
t_day = t_day.rjust(2,'0')
# リネーム後の作業日付(年月日)を取得する
today_f = t_year + t_month + t_day
# リネーム前の日経検索結果のファイル名(テンプレートファイル)
b_export_file = "【日経】検索結果_yyyymmdd.xlsx"
#　リネーム後の日経検索結果のファイル名
export_file = "【日経】検索結果_" + today_f + ".xlsx"
#  入力ファイル名
input_file = "result2.csv"

# ファイル名のリネーム処理
# ファイルをリネームします。テンプレートファイルをダウンロードしない場合、例外処理で
# 処理を終了する。
try:
    shutil.copy2(b_export_file,export_file)
except FileNotFoundError:
    print("リネーム前のファイルが存在しません")
    sys.exit()
except PermissionError:
    print("【日経】検索結果のエクセルファイルが開いています。閉じてください") 
    sys.exit()   

# 出力先のエクセルファイルをロードする
wb = op.load_workbook(export_file)
sh_name = '検索結果'
ws = wb[sh_name]
max_row = 4

# エクセルファイルに企業名、URL、データ登録日、ハイパーリンクURLを1行ずつ書き込む
with open(input_file) as f1:
    for line1 in f1:
        array1 = line1.split(",")
        company = array1[0]
        url = array1[1]
        u_date = array1[2]
        ws["F" + str(max_row)] = "link"
        ws.cell(row=max_row,column=2).value = company
        ws.cell(row=max_row,column=3).value = url
        ws.cell(row=max_row,column=4).value = u_date
        ws.cell(row=max_row,column=6).value = url
        ws.cell(row=max_row,column=6).hyperlink = url
        ws.cell(row=max_row,column=8).value = '●'
        max_row += 1
# エクセルファイルの保存
wb.save(export_file)   

