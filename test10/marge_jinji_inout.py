#######   人事異動、データのマージ処理    ######
#######   新規作成   2024/1/17         #####################
#######   作成者     takao.hattori     #####################
#######   クラスファイルの用途
#######   人事異動情報のデータを１つのファイルにマージする
#######
#######   クラスファイルの使い方        
#######   (1) 下記のコードを記載して、クラスファイルを読み込む
#######   　　from marge_jinji_inout import Marge_jinji_output
#######   (2)　インスタンスを読み込む
#######       <1> インスタンスを呼び出す
#######       r_marge = Marge_jinji_output(work_file_list,export_file,startnum,jinji_max_count)
#######       <2> データのマージ処理
#######       r_marge.main()
#######       
############################################################


# ライブラリの呼び出し
import openpyxl

class Marge_jinji_output():
    def __init__(self,work_file_list,export_file,startnum,jinji_max_count):
        self.work_file_list = work_file_list
        self.export_file = export_file
        self.startnum = startnum
        self.jinji_max_count = jinji_max_count

# マージ対象ファイルを読み出し、マージ先ファイルにコピーする。
    def main(self):  
        #マージ先のシート名を定義する      
        sheet_name = "人事異動"
        # 人事情報のマージ元一覧ファイルを読み込む
        with open(self.work_file_list,mode="r",encoding="utf-8") as f2:
            while True:
                filename2 = f2.readline()
                filename2 = filename2.replace("\n","")
                # 作業用ファイル一覧ファイルのデータがNULLの時、ループを抜ける。
                if filename2 == "":
                    break
                #マージデータ出力先のエクセルファイルを変数に定義する    
                wb = openpyxl.load_workbook(filename2)
                ws = wb[sheet_name]
                # マージ先の出力ファイルを定義する
                output1 = openpyxl.load_workbook(self.export_file)
                op = output1[sheet_name]
                # マージ対象の作業ファイルから転記したデータをマージ先ファイルにコピーする。
                for row in ws.iter_rows(min_row = self.startnum):
                    company_name = row[2].value
                    # 会社名がNULLの時にループを抜ける
                    if company_name is None:
                        break
                    changes_day = row[3].value
                    t_name = row[4].value
                    changes_item = row[5].value
                    new_div = row[6].value
                    old_div = row[7].value
                    release_day = row[8].value
                    input_day = row[9].value
                    const_sites = row[10].value
                    const_url = row[11].value
                    appending = row[12].value
                    op.cell(row=self.jinji_max_count,column=3).value = company_name
                    op.cell(row=self.jinji_max_count,column=4).value = changes_day
                    op.cell(row=self.jinji_max_count,column=5).value = t_name
                    op.cell(row=self.jinji_max_count,column=6).value = changes_item
                    op.cell(row=self.jinji_max_count,column=7).value = new_div
                    op.cell(row=self.jinji_max_count,column=8).value = old_div
                    op.cell(row=self.jinji_max_count,column=9).value = release_day
                    op.cell(row=self.jinji_max_count,column=10).value = input_day
                    op.cell(row=self.jinji_max_count,column=11).value = const_sites
                    op.cell(row=self.jinji_max_count,column=12).value = const_url
                    op.cell(row=self.jinji_max_count,column=12).hyperlink = const_url
                    op.cell(row=self.jinji_max_count,column=13).value = appending
                    # 出力先の行数を1加算する。
                    self.jinji_max_count += 1
                # マージ先エクセルファイルを保存する。    
                output1.save(self.export_file)
