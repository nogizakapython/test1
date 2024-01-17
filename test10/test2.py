#######   テスト1    ######
import os
import sys
import openpyxl
import re


list1 = os.listdir()
out_file = "test2.txt"
start_row = 0
export_file = "PRD JP_人事異動・組織改編情報List_yyyymmdd.xlsm"


def read_start():
    values1 = 5
    return values1



jinji_max_count = 5

result1 = os.path.isfile(out_file)
if result1:
    os.remove(out_file)

with open(out_file,mode="a",encoding="utf-8") as f:
    for filename in list1:
        result2 = re.match('（',filename)
        if result2:
            f.write(filename + "\n")


with open(out_file,mode="r",encoding="utf-8") as f2:
    filename2 = f2.readline()
    filename2 = filename2.replace("\n","")
    print(filename2)
    wb = openpyxl.load_workbook(filename2)
    ws = wb["人事異動"]
    startnum = read_start()
    #ob = openpyxl.load_workbook(export_file)
    #oc = ob["人事異動"]
    for row in ws.iter_rows(min_row = startnum):
        print(row[2].value)
        