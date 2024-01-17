#######   テスト1    ######
import os
import sys
import openpyxl
import re
from datetime import datetime
import shutil

dt = datetime.now()
list1 = os.listdir()
out_file = "test1list.txt"
start_row = 0
date1 = dt.strftime('%Y%m%d')
b_export_file = "PRD JP_人事異動・組織改編情報List_yyyymmdd.xlsx"
export_file = "PRD JP_人事異動・組織改編情報List_" + date1 + ".xlsx"


def read_start():
    values1 = 5
    return values1



jinji_max_count = 5

result1 = os.path.isfile(out_file)
if result1:
    os.remove(out_file)

shutil.copy(b_export_file,export_file)    

with open(out_file,mode="a",encoding="utf-8") as f:
    for filename in list1:
        result2 = re.match('（',filename)
        if result2:
            f.write(filename + "\n")


with open(out_file,mode="r",encoding="utf-8") as f2:
    while True:
        filename2 = f2.readline()
        filename2 = filename2.replace("\n","")
        print(filename2)
        if filename2 == '':
            break

        wb = openpyxl.load_workbook(filename2)
        ws = wb["人事異動"]

        output1 = openpyxl.load_workbook(export_file)
        op = output1["人事異動"]
        startnum = read_start()
        for row in ws.iter_rows(min_row = startnum):
            company_name = row[2].value
            if not company_name:
                break
            changes_day = row[3].value
            t_name = row[4].value
            changes_item = row[5].value
            new_div = row[6].value
            old_div = row[7].value
            release_day = row[8].value
            input_day = row[9].value
            const_sites = row[10].value
            const_url = row[11].value
            appending = row[12].value
            op.cell(row=jinji_max_count,column=3).value = company_name
            op.cell(row=jinji_max_count,column=4).value = changes_day
            op.cell(row=jinji_max_count,column=5).value = t_name
            op.cell(row=jinji_max_count,column=6).value = changes_item
            op.cell(row=jinji_max_count,column=7).value = new_div
            op.cell(row=jinji_max_count,column=8).value = old_div
            op.cell(row=jinji_max_count,column=9).value = release_day
            op.cell(row=jinji_max_count,column=10).value = input_day
            op.cell(row=jinji_max_count,column=11).value = const_sites
            op.cell(row=jinji_max_count,column=12).value = const_url
            op.cell(row=jinji_max_count,column=12).hyperlink = const_url
            op.cell(row=jinji_max_count,column=13).value = appending
            jinji_max_count += 1
        output1.save(export_file)


               

