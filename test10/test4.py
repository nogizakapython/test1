#######   組織改編シートのマージ処理    ######
#######   新規作成   2024/1/16         ######
#######   作成者     takao.hattori     ######
#############################################


import os
import sys
import openpyxl
import re
from datetime import datetime
import shutil

dt = datetime.now()
list1 = os.listdir()
out_file = "test4.txt"
start_row = 0
date1 = dt.strftime('%Y%m%d')
#b_export_file = "PRD JP_人事異動・組織改編情報List_yyyymmdd.xlsx"
export_file = "PRD JP_人事異動・組織改編情報List_" + date1 + ".xlsx"


def read_start():
    values1 = 6
    return values1



soshiki_max_count = 6

result1 = os.path.isfile(out_file)
if result1:
    os.remove(out_file)

    

with open(out_file,mode="a",encoding="utf-8") as f:
    for filename in list1:
        result2 = re.match('（',filename)
        if result2:
            f.write(filename + "\n")


with open(out_file,mode="r",encoding="utf-8") as f2:
    while True:
        filename2 = f2.readline()
        filename2 = filename2.replace("\n","")
        print(filename2)
        if filename2 == '':
            break

        wb = openpyxl.load_workbook(filename2)
        ws = wb["組織改編"]

        output1 = openpyxl.load_workbook(export_file)
        op = output1["組織改編"]
        startnum = read_start()
        for row in ws.iter_rows(min_row = startnum):
            company_name = row[2].value
            print(company_name)
            if company_name is None:
                break
            if company_name == '=C4':
                continue
            changes_day = row[3].value
            detail = row[4].value
            release_day = row[5].value
            input_day = row[6].value
            const_sites = row[7].value
            const_url = row[8].value
            appending = row[9].value
            op.cell(row=soshiki_max_count,column=3).value = company_name
            op.cell(row=soshiki_max_count,column=4).value = changes_day
            op.cell(row=soshiki_max_count,column=5).value = detail
            op.cell(row=soshiki_max_count,column=6).value = release_day
            op.cell(row=soshiki_max_count,column=7).value = input_day
            op.cell(row=soshiki_max_count,column=8).value = const_sites
            op.cell(row=soshiki_max_count,column=9).value = const_url
            op.cell(row=soshiki_max_count,column=9).hyperlink = const_url
            op.cell(row=soshiki_max_count,column=10).value = appending
            soshiki_max_count += 1
        output1.save(export_file)


               

