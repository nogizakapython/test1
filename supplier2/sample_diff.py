############ 一定期間のデータを抽出する処理(Sample) ###########
############ 新規作成  2024/6/13      ######################################
############ 作成者    takao.hattori  ######################################
############################################################################

# 前回の作業日を入力する関数
def data_input():
    #　前回の作業入力ルーチン
    
    ymd = ""
    while True:
        corect_count = 0
        print("前回の作業日付を西暦4桁、月2桁、日付2桁で入力してください。例) 2024年5月8日の場合は「20240508」と入力してください\n")
        data1 = input()
        if len(data1) == 8:
            corect_count += 1
        else:
            print("入力文字数が違います。")
        
        w_year = data1[:4]
        w_month = data1[4:6]
        w_day = data1[6:]
        w_year1 = int(w_year)
        w_month1 = int(w_month)
        w_day1 = int(w_day)
        # 月判定
        if w_month1 < 1 or w_month1 > 12:
            print("NG")
            
        else:     
            corect_count += 1
        # 日判定
        if w_month1 == 1 or w_month1 == 3 or w_month1 == 5 or w_month1 == 7 \
        or w_month1 == 8 or w_month1 == 10 or w_month1 == 12:
            if w_day1 <= 31:
                print("OK")
                corect_count += 1
            else:    
                print("NG")
        elif  w_month1 == 4 or w_month1 == 6 or w_month1 == 9 or w_month1 == 11:
            if w_day1 <= 30:
                print("OK")
                corect_count += 1
            else:    
                print("NG")       
        else:
            # うるう年判定
            if w_year1 % 4 == 0 or (w_year1 % 4 == 0 and w_year1 % 400 == 0):
                if w_day1 <= 29:
                    print("OK")
                    corect_count += 1
                else:
                    print("NG") 
            else:
                if w_day1 <= 28:
                    print("OK")
                    corect_count += 1                     
                else:
                    print("NG")
        if corect_count == 3:
            ymd = w_year + '/' + w_month + '/' + w_day
            break 
    return ymd        

# テンプレートから更新結果ファイルのコピー関数
def copy_excel_file():
    from datetime import datetime
    import shutil
    dt = datetime.now()
    date1 = dt.strftime('%Y%m%d')
    # 更新ファイルの定義
    base_file = "スズキ中計更新確認結果_テンプレート.xlsx"
    update_file = "スズキ中計更新確認結果_" + str(date1) + ".xlsx"
    shutil.copy(base_file,update_file)
    return update_file

  
# 前回納品日に転記した最終日と今回納品分の転記最終日を更新エクセルファイルに出力する関数
def finish_write_day(file_name,before_ymd,now_ymd):
    import openpyxl as op
    wb = op.load_workbook(file_name)
    sh_name = '更新データ'
    ws = wb[sh_name]
    ws.cell(row=2,column=4).value = before_ymd
    ws.cell(row=3,column=4).value = now_ymd
    wb.save(file_name)  


# アップデートエクセルファイルの書き込み開始行を取得する関数
def update_excel_maxcount():
    update_row_count = 6
    return update_row_count    

# 更新結果を更新結果ファイルに出力する関数(Excel)
def update_data_output(file_name,row_count,ad_company,ad_ymd,ad_title,ad_url):
    import openpyxl as op
    from openpyxl.styles.fonts import Font
    wb = op.load_workbook(file_name)
    sh_name = '更新データ'
    ws = wb[sh_name]
    ws.cell(row=row_count,column=3).value = ad_company
    ws.cell(row=row_count,column=4).value = ad_ymd
    ws.cell(row=row_count,column=5).value = ad_title
    ws.cell(row=row_count,column=5).hyperlink = ad_url
    ws.cell(row=row_count,column=5).font = Font(color='0000FF',underline='single',name='Meiryo UI')
    wb.save(file_name)    
# メイン関数
def main():
    # ライブラリのインポート
    import openpyxl as op
    from datetime import datetime
    
    from datetime import timedelta
    
    # 日付の定義
    dt = datetime.now()
    date1 = dt.strftime('%Y%m%d')
    date2 = dt.strftime('%Y/%m/%d')
    date3 = datetime.strptime(date2, '%Y/%m/%d')
    date4 = date3 - timedelta(1)
    # 前回作業日と今回作業日前日の日付を更新結果ファイルに出力する。
    update_file = copy_excel_file()
    
    # 読み込み元のファイルの変数定義
    input_file = "【IR】suzuki検索結果_" + date1 + ".xlsx"
    wb2 = op.load_workbook(input_file)
    sheet_array1 = ["スズキ"]

    # 前回自動化ツール実行日入力関数を呼び出す
    str_ymd = data_input()


    # 前回の自動化ツール実行日から、今回自動化ツール実行前日までの日付を結果ファイルに出力する
    ymd = datetime.strptime(str_ymd, '%Y/%m/%d')
    finish_write_day(update_file,ymd,date4)
    # 前回自動化ツール実行日から、今回自動化ツール実行前日までの日付を計算する。
    array1 = []
    dt1 = datetime.now()
    day_sabun = dt1 - ymd
    day_sabun = str(day_sabun)
    sabun_array1 = day_sabun.split(" ")
    day_sa = sabun_array1[0]
    day_sa = int(day_sa)

    # 前回の自動化ツール実行日から今回の自動化ツール実行前日までを日付配列に格納する。
    for i in range(0,day_sa):
        w_ymd = ymd + timedelta(i)
        w_ymd = w_ymd.strftime("%Y/%m/%d")
        array1.append(w_ymd)
    

    # 更新結果ファイルの書き込み開始行を取得する
    w1_row_count = update_excel_maxcount()

    # suzuki、IRニュース結果ファイルを読み込む 
    for sh_name in list(sheet_array1):
        ws2 = wb2[sh_name]
        w2_row_count = 5
    
        # suzuki、IRニュース結果シートの最終行を取得する
        while True:
            data_value = ws2.cell(row=w2_row_count,column=2).value
            if data_value == None:
                break
            else:
                w2_row_count += 1
        #管理用シートに書き込むデータをニュースリリース日と前回の作業日の翌日から今日までの日付で
        #突き合わせして、合致したものを管理用エクセルファイルに書き込む  
    
        for row_count in range(w2_row_count-1,4,-1):
            ad_company = sh_name
            ad_title = ws2.cell(row=row_count,column=2).value
            ad_url = ws2.cell(row=row_count,column=3).value
            ad_ymd = ws2.cell(row=row_count,column=4).value
        
            for t_ymd in array1:
                if ad_ymd == t_ymd:
                    update_data_output(update_file,w1_row_count,ad_company,ad_ymd,ad_title,ad_url)
                    w1_row_count += 1
                   
if __name__ == "__main__":
    main()    
    