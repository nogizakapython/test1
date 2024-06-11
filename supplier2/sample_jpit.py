########################################################
####   JPiTニュースリリーススクレイピング処理         ####
####   Create date  2024/6/10                      #####
####                                               #####
########################################################

# 前回出力した抽出用テキストファイルの削除関数
def remove_file(file_name):
    import os
    exist_file = os.path.isfile(file_name)
    if exist_file:
        os.remove(file_name) 
    

# タグ出力配列を作成する
def scrapping():
    import requests
    from bs4 import BeautifulSoup
    # スクレイピング対象の URL にリクエストを送り HTML を取得する
    res = requests.get('https://www.jp-it.jp/news/company/index.html')

    # レスポンスの HTML から BeautifulSoup オブジェクトを作る
    soup = BeautifulSoup(res.text, 'html.parser')
    return soup

# 日付タグ抽出関数
def span_ymd_array():
    # spanタグのタグクラス名「date」の文字列を取得する
    soup = scrapping()
    ymd_array = soup.find_all('span','date')
    return ymd_array

# URLタグ抽出関数
def td_url_array():
    # spanタグのタグクラス名「date」の文字列を取得する
    soup = scrapping()
    url_array = soup.find_all('td','news-title')
    return url_array

# タグ抽出データをテキストファイルに出力する関数
def output_text_file(line1,file_name):
    import codecs
    print(line1,file=codecs.open(file_name,'a','utf-8'))

# テンプレートファイルからコピー処理
def copyexcelfile():
    import shutil
    import datetime
    import sys
    dt = datetime.datetime.now()
    date1 = dt.strftime('%Y%m%d')    
    base_file = "JPiTテンプレート.xlsx"
    export_file = "JPiTニュース一覧" + date1 + ".xlsx"
    # Excelファイルを開いていたら、Excelを閉じる
    try:
        shutil.copy(base_file,export_file)
    except:
        print(f"{export_file}を閉じてください") 
        sys.exit()   
    return export_file

# 年月日データクレンジング処理関数
def ymd_cleansing(str1):
    ymd_array = str1.split(">")
    w_ymd = ymd_array[1]
    w_ymd = w_ymd.replace('</span','')
    w_ymd = w_ymd.replace('.','/')
    return w_ymd

# URLクレンジング処理関数
def url_cleansing(str1):
    import re
    base_url1 = "https://www.jp-it.jp"
    base_url2 = "https://www.jp-it.jp/news"
    target_url = ""
    url_array = str1.split(">")
    w_url = url_array[0]
    w_url = w_url.replace('<a href=','')
    w_url = w_url.replace('"','')
    url_check1 = re.match('../',w_url)
    # URLの相対パスを絶対パスに変換する
    if url_check1:
        url_former_str = w_url[:5]
        if url_former_str == "../..":
            w_url = w_url.replace('../..','')
            target_url = base_url1 + w_url
        else:    
            target_url = base_url2 + w_url
            target_url = target_url.replace('news..','news')
    return target_url       

# ニュースタイトルクレンジング処理関数
def title_cleansing(str1):
    import re
    title_array = str1.split(">")
    w_title = title_array[1]
    w_title = w_title.replace('</a','')
    return w_title

# Excelファイル出力関数
def outputfile(file_name,title,url,ymd,count_row):
    # ライブラリのインポート
    import openpyxl as op
    from openpyxl.styles.fonts import Font
    # Excelファイルに関数定義
    # Workbookの変数定義(ファイル名)
    wb = op.load_workbook(file_name)
    # 出力先シート名の定義
    sh_name = 'JPiT'
    # ワークシート変数定義
    ws = wb[sh_name]
    # Excelに各列の項目を出力する。
    ws.cell(row=count_row,column=2).value = title
    ws.cell(row=count_row,column=3).value = url
    ws.cell(row=count_row,column=4).value = ymd
    ws.cell(row=count_row,column=6).value = url
    ws.cell(row=count_row,column=6).hyperlink = url
    ws.cell(row=count_row,column=6).font = Font(color='0000FF',underline='single')         
    # エクセルファイルの保存
    wb.save(file_name)


# メイン関数
def main():
    import re
    # タグ抽出結果テキストファイル変数  
    output_file = "result.txt"
    max_row = 5
    
    # 前回ファイルチェックと削除処理
    remove_file(output_file) 
    # spanタグのクラス名「date」の文字列を取得する
    ymd_str = span_ymd_array()
    # tdタグのクラス名「news-title」の文字列を取得する
    url_str = td_url_array()
    
    # spanタグの配列の要素数を取得する
    length1 = len(ymd_str)
    # tdタグの配列の要素数を取得する
    length2 = len(url_str)
    # spanタグの配列数とtdタグの配列数が同じとき、spanタグ、tdタグの順番にタグを抽出し、
    # タグファイルをテキストファイルに出力する
    if length1 == length2:
        for i in range(length1):
            output_text_file(ymd_str[i],output_file)
            output_text_file(url_str[i],output_file)

    # テンプレートExcelファイルからスクレイピング用Excelファイルをコピーする
    export_file = copyexcelfile()

    # 抽出したタグファイルを読み込んでエクセルファイルに書き込む
    fileobj = open(output_file,encoding="utf-8")
    while True:
        line1 = fileobj.readline()
        line1 = line1.replace("\n","")
        if not line1:
            break
           
        result1 = re.search('date',line1)
        result2 = re.search('<a href',line1)
        if result1:
            w_ymd = ymd_cleansing(line1) 
            
        if result2:
            w_url = url_cleansing(line1)
            w_title = title_cleansing(line1)
            outputfile(export_file,w_title,w_url,w_ymd,max_row)
            max_row += 1

           
if __name__ == "__main__":
    main()
