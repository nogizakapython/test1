#######   スズキ スクレイピングサンプル　###########
#######   新規作成  2024/06/11  ##########
#######   Author  takao.hattori ###########


# 前回実行用のタグ抽出ファイルを削除する。
def remove_before_file():
    out_file = "suzuki.txt"
    import os
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)
    return out_file    

# 結果Excelファイルをテンプレートファイルからコピーする関数
def make_result_file():
    import shutil
    import datetime
    dt = datetime.datetime.now()
    import sys
    date1 = dt.strftime('%Y%m%d')   
    base_file = "【IR】suzuki検索結果_テンプレート.xlsx"
    export_file = "【IR】suzuki検索結果_" + date1 + ".xlsx"
    try:
        shutil.copy(base_file,export_file)
    except:
        print(f"{export_file}が開いています。閉じて再実行してください。")
        sys.exit()    
    return export_file

# スクレイピング処理関数
def scrapping(output_file):
    # WebDriverライブラリをインポート
    from selenium import webdriver
    from time import sleep
    import selenium
    from selenium.webdriver.common.by import By
    import codecs
    import sys
    # スクレイピング先URL変数の定義
    target_url = 'https://www.suzuki.co.jp/ir/news/'
    # WebDriverの定義(Chrome Driverを使う)
    driver = webdriver.Chrome()
    # Chromeを開いて企業HPにアクセスし、5秒間待機する
    driver.get(target_url)
    sleep(5)
    # 外側の例外処理はxpathが違うときに異常終了する。
    try:    
        for i in range(1,6):
            for j in range(1,31):
                # 内側の例外処理は既定の件数に満たないときにループから抜ける
                try:
                    xpath_str1 = '/html/body/div[4]/div[5]/div/section[' + str(i) + ']/dl/a[' + str(j) + ']'
                    element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                    print(element_str1.get_attribute("outerHTML"),file=codecs.open(output_file,'a','utf-8'))
                except:
                    break
    except selenium.common.exceptions.NoSuchElementException as e:
        sys.exit()


    # 画面を閉じる
    driver.quit()

# URLクレンジング処理関数
def url_clansing(str1):
    import re
    base_url = 'https://www.suzuki.co.jp'

    w_array1 = str1.split(">")
    w_urlstr = w_array1[0]
    w_array2 = w_urlstr.split("=")
    w_urlstr = w_array2[1]
    w_urlstr = w_urlstr.replace('"','')
    w_urlstr = w_urlstr.replace('target','')
    w_url = base_url + w_urlstr  
    return w_url

# リリース日クレンジング処理関数
def ymd_clansing(str1):
    w_array3 = str1.split(">")
    w_ymd_str = w_array3[2]
    w_ymd_str = w_ymd_str.replace("</time",'')
    w_ymd_str = w_ymd_str.replace("年",'/')
    w_ymd_str = w_ymd_str.replace("月",'/')
    w_ymd_str = w_ymd_str.replace("日",'')
    w_ymd_str = w_ymd_str.replace("<",'')
    w_ymd_str = w_ymd_str.replace('span class="news_pdf"',"")
    ymd_array = w_ymd_str.split("/")
    ymd_year = ymd_array[0]
    ymd_month = int(ymd_array[1])
    ymd_day = int(ymd_array[2])
    if ymd_month < 10:
        ymd_month = "0" + str(ymd_month)
    else:
        ymd_month = str(ymd_month)
    if ymd_day < 10:
        ymd_day = "0" + str(ymd_day)
    else:
        ymd_day = str(ymd_day)
    w_ymd = ymd_year + "/" + ymd_month + "/" + ymd_day

    return w_ymd

# タイトル名クレンジング処理関数
def title_clansing(str1):
    w_array4 = str1.split(">")
    w_title_str = w_array4[4]
    if w_title_str == "</span":
        w_title_str = w_array4[7]
    w_title = w_title_str.replace("</dd",'')
    
    return w_title

# 検索ワードにヒットしたタイトル名、URL、発表日を行数を指定してExcelファイルに出力する関数。
def result_output_file(file_name,row_count,w_title,w_url,w_ymd):
     import openpyxl as op
     from openpyxl.styles.fonts import Font
     wb = op.load_workbook(file_name)
     sh_name = 'スズキ'
     ws = wb[sh_name]
     ws.cell(row=row_count,column=2).value = w_title
     ws.cell(row=row_count,column=3).value = w_url
     ws.cell(row=row_count,column=4).value = w_ymd
     ws.cell(row=row_count,column=6).value = w_url
     ws.cell(row=row_count,column=6).hyperlink = w_url
     ws.cell(row=row_count,column=6).font = Font(color='0000FF',underline='single')
     # エクセルファイルの保存
     wb.save(file_name)

# メイン処理関数
def main():
    # ライブラリをインポート
    import re
    # Excel出力先の行数変数
    max_row = 5
    # Excelに出力するタイトル名の検索ワード指定変数
    key_word = r"(決算|報告書|経営)"

    out_file = remove_before_file()
    # 時間を計るライブラリをインポート
    export_file = make_result_file()    
    scrapping(out_file)
    fileobj = open(out_file,encoding="utf-8")
    while True:
        line1 = fileobj.readline()
        line1 = line1.replace("\n","")
        if not line1:
            break
      
        result1 = re.match('<a href',line1)
    
    
        if result1:
            w_url = url_clansing(line1)
            w_ymd = ymd_clansing(line1)
            w_title = title_clansing(line1)
            title_result = re.search(key_word,w_title)
            if title_result:
               result_output_file(export_file,max_row,w_title,w_url,w_ymd)    
               max_row += 1
                
if __name__ == "__main__":
    main()
