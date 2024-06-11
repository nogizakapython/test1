#######   電通総研 スクレイピングサンプル　###########
#######   新規作成  2024/06/11  ##########
#######   Author  takao.hattori ###########


# 前回実行用のタグ抽出ファイルを削除する。
def remove_before_file():
    out_file = "dentsusoken.txt"
    import os
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)
    return out_file    

# 結果Excelファイルをテンプレートファイルからコピーする関数
def make_result_file():
    import shutil
    import datetime
    dt = datetime.datetime.now()
    import sys
    date1 = dt.strftime('%Y%m%d')   
    base_file = "DSI人事テンプレート.xlsx"
    export_file = "DSI人事" + date1 + ".xlsx"
    try:
        shutil.copy(base_file,export_file)
    except:
        print(f"{export_file}が開いています。閉じて再実行してください。")
        sys.exit()    
    return export_file

# スクレイピング処理関数
def scrapping(output_file):
    # WebDriverライブラリをインポート
    from selenium import webdriver
    from time import sleep
    import selenium
    from selenium.webdriver.common.by import By
    import codecs
    import sys
    # スクレイピング先URL変数の定義
    target_url = 'https://www.dentsusoken.com/news'
    # WebDriverの定義(Chrome Driverを使う)
    driver = webdriver.Chrome()
    # Chromeを開いて企業HPにアクセスし、5秒間待機する
    driver.get(target_url)
    sleep(5)
    # 外側の例外処理はxpathが違うときに異常終了する。
    try:    
        for i in range(1,13):
            for j in range(1,21):
                # 内側の例外処理は既定の件数に満たないときにループから抜ける
                try:
                    xpath_str1 = '/html/body/div[1]/main/div/div[2]/div[3]/div/div/div/ul[' + str(i) + ']/li[' + str(j) + ']/a'
                    element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                    print(element_str1.get_attribute("outerHTML"),file=codecs.open(output_file,'a','utf-8'))
                except:
                    break
    except selenium.common.exceptions.NoSuchElementException as e:
        sys.exit()


    # 画面を閉じる
    driver.quit()

# URLクレンジング処理関数
def url_clansing(str1):
    import re
    base_url = 'https://www.dentsusoken.com'

    w_array1 = str1.split("=")
    w_urlstr = w_array1[2]
    w_urlstr = w_urlstr.replace('"','')
    w_urlstr = w_urlstr.replace('>','')
    w_urlstr = w_urlstr.replace(' ','')
    w_urlstr = w_urlstr.replace('target','')
    url_check = re.search('pdf',w_urlstr)
    if url_check:
        w_url = 'https:' + w_urlstr
    else:
        w_url = base_url + w_urlstr  
    return w_url

# リリース日クレンジング処理関数
def ymd_clansing(str1):
    w_array2 = str1.split(">")
    w_ymd_str = w_array2[1]
    w_ymd_str = w_ymd_str.replace("</time",'')
    w_ymd_str = w_ymd_str.replace("年",'/')
    w_ymd_str = w_ymd_str.replace("月",'/')
    w_ymd = w_ymd_str.replace("日",'')
    return w_ymd

# タイトル名クレンジング処理関数
def title_clansing(str1):
    w_array3 = str1.split(">")
    w_title_str = w_array3[1]
    w_title = w_title_str.replace("</span",'')
    return w_title

# 検索ワードにヒットしたタイトル名、URL、発表日を行数を指定してExcelファイルに出力する関数。
def result_output_file(file_name,row_count,w_title,w_url,w_ymd):
     import openpyxl as op
     from openpyxl.styles.fonts import Font
     wb = op.load_workbook(file_name)
     sh_name = '電通総研'
     ws = wb[sh_name]
     ws.cell(row=row_count,column=2).value = w_title
     ws.cell(row=row_count,column=3).value = w_url
     ws.cell(row=row_count,column=4).value = w_ymd
     ws.cell(row=row_count,column=6).value = w_url
     ws.cell(row=row_count,column=6).hyperlink = w_url
     ws.cell(row=row_count,column=6).font = Font(color='0000FF',underline='single')
     # エクセルファイルの保存
     wb.save(file_name)

# メイン処理関数
def main():
    # ライブラリをインポート
    import re
    # Excel出力先の行数変数
    max_row = 5
    # Excelに出力するタイトル名の検索ワード指定変数
    key_word = r"(人事異動|役員|就任)"

    out_file = remove_before_file()
    # 時間を計るライブラリをインポート
    export_file = make_result_file()    
    scrapping(out_file)
    fileobj = open(out_file,encoding="utf-8")
    while True:
        line1 = fileobj.readline()
        line1 = line1.replace("\n","")
        if not line1:
            break
           
        result1 = re.search('mod-list-news__link',line1)
        result2 = re.search('mod-list-news__date',line1)
        result3 = re.search('mod-list-news__text',line1)
    
        if result1:
            w_url = url_clansing(line1)
    
        if result2:
            w_ymd = ymd_clansing(line1)   
        
        if result3:
            w_title = title_clansing(line1)
            title_result = re.search(key_word,w_title)
            if title_result:
                result_output_file(export_file,max_row,w_title,w_url,w_ymd)    
                max_row += 1
                
if __name__ == "__main__":
    main()
