#######   IC Energy コスモ石油　###########
#######   新規作成  2024/2/28  ##########
#######   修正      2024/4/11 ２社にまたがって同じニュースがある場合は両方出力に変更
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from datetime import datetime , timedelta


dt = datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
# date2 = 2023
date4 = dt.strftime('%Y%m%d')
date5 = int(date2) - 1
date6 = int(date2) - 2

# 今日の日付を取得
today1 = dt.strftime('%Y/%m/%d')
# today1 = "2023/04/07"
# 今日の曜日を取得
today_day = dt.strftime('%a')
# 昨日の日付を取得
yesterday = dt - timedelta(1)
yesterday1 = yesterday.strftime('%Y/%m/%d')

file_name = "IC_cosmo" + date1 + ".txt"
out_file = "IC_cosmo.txt"
base_url = 'https://www.cosmo-energy.co.jp'
access_url1 = 'https://www.cosmo-energy.co.jp/ja/about/press.html'
access_url2 = 'https://www.cosmo-energy.co.jp/ja/about/press/'

target_url1 = 'https://www.cosmo-energy.co.jp/ja/about/press.html'
target_url2 = ""
#############################################################################
#  list id変数定義(年に1回、毎年4月に新年度のデータが入ってきたら変数を更新する)
#  id_list2の変数にはid_list1の変数を変える。id_list1の変数に新しいid名を定義する
id_list1 = "list-125886634c"
id_list2 = "list-44326da321"
#############################################################################
max_row = 0
base_file = "ICEnergyニュースリリース一覧テンプレート.xlsx"
export_file = "ICEnergyニュースリリース出力用" + date4 + ".xlsx"

# 企業名配列の定義
company_array = []
# 決算変数の定義
output_company_name = ""

# write_flag = 0
xpath_str1 = ""


# 年度でニュースリリースが変わるので、年度判定処理
def web_scrapping():

    w_ymd_array = today1.split('/')
    w_month = int(w_ymd_array[1])
    w_day = int(w_ymd_array[2])
    if w_month < 4:
        target_url2 = access_url2 + str(date6) + '.html'
    else:
        target_url2 = access_url2 + str(date5) + '.html'    

        
    # Chromeを指定する
    driver = webdriver.Chrome()
    
    for url in [target_url1,target_url2]:
        # Chromeを開いて企業検索にアクセスする
        driver.get(url)
        sleep(3)

        # 最新15件のニュースリリースを取得

        for i in range(1,16):
            try:
                if url == access_url1:
                    
                    xpath_str1 = '//*[@id="' + id_list1 + '"]/li[' + str(i) + ']'
                    
                else:        
                    xpath_str1 = '//*[@id="' + id_list2 + '"]/li[' +str(i) + ']'
                        
                    
                element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
            except:
                if i == 1:
                    print("データがありません",file=codecs.open(file_name,'a','utf-8')) 
    # 画面を閉じる
    driver.quit()


def copy_excel_file():
    # ニュースリリース作業用エクセルファイルがなければ作成、あればそのまま
    
    file_exist = os.path.isfile(export_file)
    if file_exist:
        print("作業用エクセルファイル作成済です。")
    else:
        shutil.copy2(base_file,export_file)

def log_lotate():
    # 前回実行の作業ファイルがあれば削除
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)

    # 作業用ログファイルをコピー
    shutil.copy2(file_name,out_file)

def get_max_row():
    w_row_count = 4
    while True:
        cell_value = ws.cell(row=w_row_count,column=3).value
        # print(cell_value)
        if cell_value == None:
            break
        else:
            w_row_count += 1
            # print(w_row_count)
    return w_row_count



web_scrapping()
copy_excel_file()
wb = op.load_workbook(export_file)
sh_name = 'コスモ石油'
ws = wb[sh_name]
log_lotate()
max_row = get_max_row()

# ファイルを読み込み、URL、ニュースリリース日、ニュースリリース対象企業、リンクのタイトルを取得
fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1 == "":
        break
    result1 = re.search("<a",line1)
    result2 = re.search("cmp-news-index__item-date",line1)
    result3 = re.search("cmp-news-index__item-category-name",line1)
    result4 = re.match("        </div>",line1)
    result5 = re.search("cmp-news-index__item-description",line1)

    if result1:
        w_array1 = line1.split("=")
        w_soutai_url = w_array1[4]
        w_soutai_url = w_soutai_url.replace(">","")
        w_soutai_url = w_soutai_url.replace('"','')
        w_url = base_url + w_soutai_url
        # print(w_url)
    
    if result2:
        w_array2 = line1.split(">")
        w_ymd = w_array2[1]
        w_ymd = w_ymd.replace('-',"/")
        w_ymd = w_ymd.replace('</span',"")
        # print(w_ymd)
    
    if result3:
        w_array3 = line1.split(">")  
        w_cosmo_corp_name = w_array3[1]
        w_cosmo_corp_name = w_cosmo_corp_name.replace("</span","")
        company_array.append(w_cosmo_corp_name)
    
    if result4:
        company_count = len(company_array)
        company_result1 = "お知らせ" in company_array
        if company_result1:
            company_count = 1
        
        company_result2 = "決算" in company_array
        if company_result2 or company_result2:
            company_count = 1
        
        if company_count == 1:
            output_company_name = company_array[0]
            # print(output_company_name)   
        
                    
    if result5:
        line1 = line1.replace("<br>","")
        w_array3 = line1.split(">")
        w_title = w_array3[1]
        w_title = w_title.replace("</span","")
        if company_count == 2:
            for c_name in list(company_array):
                ws.cell(row=max_row,column=3).value = c_name
                ws.cell(row=max_row,column=4).value = w_ymd
                ws.cell(row=max_row,column=5).value = w_title
                ws.cell(row=max_row,column=5).hyperlink = w_url
                ws.cell(row=max_row,column=6).value = today1
                ws.cell(row=max_row,column=7).value = w_url                
                max_row += 1
                # エクセルファイルの保存
                wb.save(export_file)
        else:
            ws.cell(row=max_row,column=3).value = output_company_name
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=5).value = w_title
            ws.cell(row=max_row,column=5).hyperlink = w_url
            ws.cell(row=max_row,column=6).value = today1
            ws.cell(row=max_row,column=7).value = w_url                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)    

        company_array = [] 
        output_company_name = ""                 

