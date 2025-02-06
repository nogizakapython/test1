#######   IC Energy 岩谷産業（IR）　###########
#######   新規作成  2024/3/29  ##########
#######   修正     2025/2/7 xpath修正  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from datetime import datetime , timedelta


dt = datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
# date2 = 2024
date3 = int(date2) - 1
date4 = dt.strftime('%Y%m%d')


# 今日の日付を取得
today1 = dt.strftime('%Y/%m/%d')
# today1 = "2024/01/11"
# 今日の曜日を取得
today_day = dt.strftime('%a')

file_name = "IC_iwatani_ir" + date1 + ".txt"
out_file = "IC_iwatani_ir.txt"
base_url = 'https://www.iwatani.co.jp'
access_url = 'https://www.iwatani.co.jp/jpn/ir/news/'
target_url = ""
target_url1 = ""
target_url2 = ""
max_row = 5
# base_file = "ICEnergyニュースリリース一覧テンプレート.xlsx"
export_file = "ICEnergyニュースリリース出力用" + date4 + ".xlsx"
# 入力用ファイルの行数変数
row_count = 0
# 企業名配列の定義
# company_array = []
# 決算変数の定義
output_company_name = ""

# write_flag = 0
xpath_str1 = ""

# 検索URL取得関数
def get_url(year):
    work_url = access_url + str(year) + "/"
    return work_url

# 年度でニュースリリースが変わるので、年度判定処理
def web_scrapping():
    w_ymd_array = today1.split('/')
    w_month = int(w_ymd_array[1])
    w_day = int(w_ymd_array[2])
    
    if w_month == 1 and w_day <= 12:
        target_url1 = get_url(date2)
        target_url2 = get_url(date3)
    else:
        target_url = get_url(date2)
    
    
    # Chromeを指定する
    driver = webdriver.Chrome()
    

    driver.get(access_url)
    sleep(3)
    # 2025/2/6 xpath変更に伴う修正 
    for i in range(1,16):
        try:                           
            xpath_str1 = '/html/body/div[2]/div[2]/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/table/tbody/tr[' + str(i) + ']'
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
            print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
                            
        except:
            if i == 1:
                print("データがありません",file=codecs.open(file_name,'a','utf-8')) 
    
    # 画面を閉じる
    driver.quit()



def log_lotate():
    # 前回実行の作業ファイルがあれば削除
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)

    # 作業用ログファイルをコピー
    shutil.copy2(file_name,out_file)

web_scrapping()
company_name = "岩谷産業IR"
wb = op.load_workbook(export_file)
sh_name = '岩谷産業IR'
ws = wb[sh_name]
log_lotate()


# ファイルを読み込み、URL、ニュースリリース日、ニュースリリース対象企業、リンクのタイトルを取得
fileobj = open(out_file,encoding="utf-8")
while True:
     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     if line1 == "" and row_count > 100:
         break
     else:
         row_count += 1
     result1 = re.match("								20",line1)
     result2 = re.search("img src",line1)

     if result1:
         w_ymd = line1.replace("								","")
        #  print(w_ymd)


     if result2:
         w_array1 = line1.split(">")
         w_data1 = w_array1[1]
         w_array2 = w_data1.split("=")
         w_soutai_url = w_array2[1]
         w_soutai_url = w_soutai_url.replace('"','')
         w_soutai_url = w_soutai_url.replace(" class","")
         w_url = base_url + w_soutai_url    
        #  print(w_url)
         w_title = w_array1[2]
         w_title = w_title.replace('<span class="icon_pdf"',"")
         w_title = w_title.replace('<span class="icon_new"',"")
         w_title = w_title.replace('</a',"")
        #  print(w_title)
 
         ws.cell(row=max_row,column=3).value = company_name
         ws.cell(row=max_row,column=4).value = w_ymd
         ws.cell(row=max_row,column=5).value = w_title
         ws.cell(row=max_row,column=5).hyperlink = w_url
         ws.cell(row=max_row,column=6).value = today1
         ws.cell(row=max_row,column=7).value = w_url
                        
         max_row += 1
         # エクセルファイルの保存
         wb.save(export_file)

