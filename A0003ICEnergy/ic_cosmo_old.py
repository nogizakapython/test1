#######   IC Energy コスモ石油　###########
#######   新規作成  2024/2/28  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from datetime import datetime , timedelta


dt = datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = int(date2) - 1
# 今日の日付を取得
today1 = dt.strftime('%Y/%m/%d')
# 昨日の日付を取得
yesterday = dt - timedelta(1)
yesterday1 = yesterday.strftime('%Y/%m/%d')

file_name = "IC_cosmo" + date1 + ".txt"
out_file = "IC_cosmo.txt"
base_url = 'https://www.cosmo-energy.co.jp'
access_url = 'https://www.cosmo-energy.co.jp/ja/about/press.html#'
target_url = ""
# max_row = 5
base_file = "ICEnergyニュースリリース一覧テンプレート.xlsx"
export_file = "ICEnergyニュースリリース出力用.xlsx"
row_count = 0
# write_flag = 0
xpath_str1 = ""
# 企業名配列の定義
company_array = []
# 決算変数の定義
output_company_name = ""

# 検索URL取得関数
def get_url(year):
    work_url = access_url + str(year)
    return work_url

# 年度でニュースリリースが変わるので、年度判定処理
w_ymd_array = today1.split('/')
w_month = int(w_ymd_array[1])
if w_month <= 3:
    target_url = get_url(date3)
else:
    target_url = get_url(date2)


# Chromeを指定する
driver = webdriver.Chrome()

# Chromeを開いて企業検索にアクセスする
driver.get(target_url)
sleep(5)

# 最新20件のニュースリリースを取得 
for i in range(1,21):
    xpath_str1 = '//*[@id="list-0c2c1d0398"]/li[' + str(i) + ']/a'
    element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
    print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        
# 画面を閉じる
driver.quit()


# ニュースリリース作業用エクセルファイルがなければ作成、あればそのまま
file_exist = os.path.isfile(export_file)
if file_exist:
   print("作業用エクセルファイル作成済です。")
else:
   shutil.copy2(base_file,export_file)

# 前回実行の作業ファイルがあれば削除
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

# 作業用ログファイルをコピー
shutil.copy2(file_name,out_file)


# ファイルを読み込み、URL、ニュースリリース日、ニュースリリース対象企業、リンクのタイトルを取得
fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1 == "":
        break
    
    
    result1 = re.match("<a",line1)
    result2 = re.search("cmp-news-index__item-date",line1)
    result3 = re.search("cmp-news-index__item-category-name",line1)
    result4 = re.match("        </div>",line1)
    result5 = re.search("cmp-news-index__item-description",line1)

    if result1:
        w_array1 = line1.split("=")
        w_soutai_url = w_array1[4]
        w_soutai_url = w_soutai_url.replace(">","")
        w_soutai_url = w_soutai_url.replace('"','')
        w_url = base_url + w_soutai_url
        # print(w_url)   
    
    if result2:
        w_array2 = line1.split(">")
        w_ymd = w_array2[1]
        w_ymd = w_ymd.replace('-',"/")
        w_ymd = w_ymd.replace('</span',"")
        # print(w_ymd)
    
    if result3:
        w_array3 = line1.split(">")  
        w_cosmo_corp_name = w_array3[1]
        w_cosmo_corp_name = w_cosmo_corp_name.replace("</span","")
        company_array.append(w_cosmo_corp_name)
    
    if result4:
        company_count = len(company_array)
        company_result1 = "お知らせ" in company_array
        if company_result1:
            company_count = 1
        
        company_result2 = "決算" in company_array
        if company_result2 or company_result2:
            company_count = 1
        
        if company_count == 2:
            for c_name in list(company_array):
                if output_company_name == "":
                    output_company_name = c_name
                elif output_company_name == "コスモエネルギーホールディングス" and not c_name == "コスモエネルギーホールディングス":
                    output_company_name = c_name
        elif company_count == 1:
            output_company_name = company_array[0]

        # print(output_company_name)   
        company_array = [] 
        output_company_name = ""
                    
    if result5:
        line1 = line1.replace("<br>","")
        w_array3 = line1.split(">")
        w_title = w_array3[1]
        w_title = w_title.replace("</span","")
        wb = op.load_workbook(export_file)
        sh_name = 'コスモ石油'
        ws = wb[sh_name]
        #  ws.cell(row=max_row,column=2).value = w_title
            # ws.cell(row=max_row,column=3).value = w_url
            # ws.cell(row=max_row,column=4).value = w_ymd
            # ws.cell(row=max_row,column=6).value = w_url
            # ws.cell(row=max_row,column=6).hyperlink = w_url
                
            # max_row += 1
            # エクセルファイルの保存
            # try:
                # wb.save(export_file)
            # except PermissionError as e:
                # fname = export_file
                # openerr = company_fileopenerror.ReadfileError(fname)
                # openerr.readerror()
                # sys.exit()               
         
