#######   IC Energy 出光興産　###########
#######   新規作成  2024/3/11  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from datetime import datetime , timedelta


dt = datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = int(date2) - 1
date4 = dt.strftime('%Y%m%d')
# 今日の日付を取得
today1 = dt.strftime('%Y/%m/%d')
# 今日の曜日を取得
today_day = dt.strftime('%a')
# 昨日の日付を取得
yesterday = dt - timedelta(1)
yesterday1 = yesterday.strftime('%Y/%m/%d')

file_name = "IC_idemitsu" + date1 + ".txt"
out_file = "IC_idemitsu.txt"
base_url = 'https://www.idemitsu.com'
access_url = 'https://www.idemitsu.com/jp/news/index.html'
target_url = 'https://www.idemitsu.com/jp/news/index.html'
max_row = 0
# base_file = "ICEnergyニュースリリース一覧テンプレート.xlsx"
export_file = "ICEnergyニュースリリース出力用" + date4 + ".xlsx"

# 企業名配列の定義
# company_array = []
# 決算変数の定義
output_company_name = ""

# write_flag = 0
xpath_str1 = ""

# 検索URL取得関数
def get_url(year):
    work_url = access_url + str(year)
    return work_url

# 年度でニュースリリースが変わるので、年度判定処理
def web_scrapping():
    

    # Chromeを指定する
    driver = webdriver.Chrome()

    # Chromeを開いて企業検索にアクセスする
    driver.get(target_url)
    sleep(3)
    button_xpath = '/html/body/main/article/div/section[4]/section[2]/a/span[1]'
    button1 = driver.find_element(By.XPATH,button_xpath)
    driver.execute_script('arguments[0].click();',button1)
    sleep(3)
    # 最新20件のニュースリリースを取得 
    for i in range(1,16):
        try:
            xpath_str1 = '/html/body/main/article/div/section[4]/section[1]/div[1]/ul/li[' + str(i) + ']'
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
            print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        except:
            if i == 1:
                print("データがありません",file=codecs.open(file_name,'a','utf-8'))   
        
    # 画面を閉じる
    driver.quit()



def log_lotate():
    # 前回実行の作業ファイルがあれば削除
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)

    # 作業用ログファイルをコピー
    shutil.copy2(file_name,out_file)

def get_max_row():
    w_row_count = 4
    while True:
        cell_value = ws.cell(row=w_row_count,column=3).value
        # print(cell_value)
        if cell_value == None:
            break
        else:
            w_row_count += 1
            # print(w_row_count)
    return w_row_count



web_scrapping()
company_name = "出光"
wb = op.load_workbook(export_file)
sh_name = '出光'
ws = wb[sh_name]
log_lotate()
max_row = get_max_row()
# ファイルを読み込み、URL、ニュースリリース日、ニュースリリース対象企業、リンクのタイトルを取得
fileobj = open(out_file,encoding="utf-8")
while True:
     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     if line1 == "":
         break
     result1 = re.match("    <a href",line1)
     result2 = re.match('            <span class',line1)
     result3 = re.match('              <p class="typography__inner',line1)
     

     if result1:
         w_array1 = line1.split("=")
         w_soutai_url = w_array1[1]
         w_soutai_url = w_soutai_url.replace('"','')
         w_soutai_url = w_soutai_url.replace(">","")
         w_soutai_url = w_soutai_url.replace(" class","")
         w_url = base_url + w_soutai_url    
        #  print(w_url)
    
     if result2:
         w_array2 = line1.split(">")
         w_ymd = w_array2[1]
         w_ymd = w_ymd.replace('</span',"")
         w_ymd = w_ymd.replace('年',"/")
         w_ymd = w_ymd.replace('月',"/")
         w_ymd = w_ymd.replace('日',"")
         ymd_array = w_ymd.split("/")
         ymd_year = ymd_array[0]
         ymd_month = int(ymd_array[1])
         ymd_day = int(ymd_array[2])
         if ymd_month < 10:
            ymd_month = "0" + str(ymd_month)
         if ymd_day < 10:
            ymd_day = "0" + str(ymd_day)
         w_ymd = str(ymd_year) + "/" + str(ymd_month) + "/" + str(ymd_day)
        #  print(w_ymd)
    
     if result3:
         w_array3 = line1.split(">")  
         w_title = w_array3[1]
         w_title = w_title.replace('</p',"")
        #  print(w_title)
    
         ws.cell(row=max_row,column=3).value = company_name
         ws.cell(row=max_row,column=4).value = w_ymd
         ws.cell(row=max_row,column=5).value = w_title
         ws.cell(row=max_row,column=5).hyperlink = w_url
         ws.cell(row=max_row,column=6).value = today1
         ws.cell(row=max_row,column=7).value = w_url
                
         max_row += 1
         # エクセルファイルの保存
         wb.save(export_file)
