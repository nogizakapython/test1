#######   IC Energy エネオス　###########
#######   新規作成  2024/3/11  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from datetime import datetime , timedelta


dt = datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
# date2 = 2023
date3 = int(date2) - 1
# date3 = 2022
date4 = dt.strftime('%Y%m%d')
# 今日の日付を取得
today1 = dt.strftime('%Y/%m/%d')
# today1 = "2023/04/07"
# 今日の曜日を取得
today_day = dt.strftime('%a')
# 昨日の日付を取得
yesterday = dt - timedelta(1)
yesterday1 = yesterday.strftime('%Y/%m/%d')

file_name = "IC_eneos" + date1 + ".txt"
out_file = "IC_eneos.txt"
base_url = 'https://www.eneos.co.jp'
access_url = 'https://www.eneos.co.jp/newsrelease/'
target_url = 'https://www.eneos.co.jp/newsrelease/'
target_url1 = ''
target_url2 = ''
max_row = 5
# base_file = "ICEnergyニュースリリース一覧テンプレート.xlsx"
export_file = "ICEnergyニュースリリース出力用" + date4 + ".xlsx"

# 企業名配列の定義
# company_array = []
# 決算変数の定義
output_company_name = ""

# write_flag = 0
xpath_str1 = ""

# 検索URL取得関数
def get_url(year):
    work_url = access_url + str(year)
    return work_url

# 年度でニュースリリースが変わるので、年度判定処理
def web_scrapping():
    w_ymd_array = today1.split('/')
    w_month = int(w_ymd_array[1])
    w_day = int(w_ymd_array[2])
    if w_month <= 3 :
        target_url = get_url(date3)
    elif w_month == 4 and w_day < 8:
        target_url1 = get_url(date2)
        target_url2 = get_url(date3)
           
    else:
        target_url = get_url(date2)

    

    # Chromeを指定する
    driver = webdriver.Chrome()

    # Chromeを開いて企業検索にアクセスする
    if w_month == 4 and w_day < 8:
        for url in [target_url1,target_url2]:
            driver.get(url)
            sleep(2)
            for i in range(1,16):
                try:
                    xpath_str1 = '//*[@id="mainBody"]/div[2]/div/section/div[2]/div[2]/div/article[' + str(i) + ']/a'
                    element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                    print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
                except:
                    if i == 1:
                        print("データがありません",file=codecs.open(file_name,'a','utf-8'))
    else:       
        driver.get(target_url)
        sleep(3)
        # 最新20件のニュースリリースを取得 
        for i in range(1,16):
            try:
                xpath_str1 = '//*[@id="mainBody"]/div[2]/div/section/div[2]/div[2]/div/article[' + str(i) + ']/a'
                element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
            except:
                if i == 1:
                    print("データがありません",file=codecs.open(file_name,'a','utf-8'))
    # 画面を閉じる
    driver.quit()



def log_lotate():
    # 前回実行の作業ファイルがあれば削除
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)

    # 作業用ログファイルをコピー
    shutil.copy2(file_name,out_file)



web_scrapping()
company_name = "エネオス"
wb = op.load_workbook(export_file)
sh_name = 'エネオス'
ws = wb[sh_name]
log_lotate()

# ファイルを読み込み、URL、ニュースリリース日、ニュースリリース対象企業、リンクのタイトルを取得
fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1 == "":
       break
    result1 = re.match("<a class",line1)
    result2 = re.search("news_date",line1)
    result3 = re.search("news_title",line1)
     

    if result1:
        w_array1 = line1.split("=")
        w_soutai_url = w_array1[2]
        w_soutai_url = w_soutai_url.replace('"','')
        w_soutai_url = w_soutai_url.replace(">","")
        w_url = base_url + w_soutai_url    
        # print(w_url)
    
    if result2:
        w_array2 = line1.split(">")
        w_ymd = w_array2[2]
        w_ymd = w_ymd.replace('</span',"")
        w_ymd = w_ymd.replace('年',"/")
        w_ymd = w_ymd.replace('月',"/")
        w_ymd = w_ymd.replace('日',"")
        w_ymd = w_ymd.replace(' ',"")
        ymd_array = w_ymd.split("/")
        ymd_year = ymd_array[0]
        ymd_month = int(ymd_array[1])
        ymd_day = int(ymd_array[2])
        if ymd_month < 10:
            ymd_month = "0" + str(ymd_month)
        if ymd_day < 10:
            ymd_day = "0" + str(ymd_day)
        w_ymd = str(ymd_year) + "/" + str(ymd_month) + "/" + str(ymd_day)
        # print(w_ymd)
    
    if result3:
        w_array3 = line1.split(">")  
        w_title = w_array3[1]
        w_title = w_title.replace('</h3',"")
        # print(w_title)
    
        ws.cell(row=max_row,column=3).value = company_name
        ws.cell(row=max_row,column=4).value = w_ymd
        ws.cell(row=max_row,column=5).value = w_title
        ws.cell(row=max_row,column=5).hyperlink = w_url
        ws.cell(row=max_row,column=6).value = today1
        ws.cell(row=max_row,column=7).value = w_url                
        max_row += 1
        # エクセルファイルの保存
        try:
           wb.save(export_file)
        except PermissionError as e:
           sys.exit()  


