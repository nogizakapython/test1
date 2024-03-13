#######   IC Energy 岩谷産業　###########
#######   新規作成  2024/3/11  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
from datetime import datetime , timedelta


dt = datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
# date2 = 2024
date3 = int(date2) - 1
date4 = dt.strftime('%Y%m%d')


# 今日の日付を取得
today1 = dt.strftime('%Y/%m/%d')
# today1 = "2024/01/11"
# 今日の曜日を取得
today_day = dt.strftime('%a')

file_name = "IC_iwatani" + date1 + ".txt"
out_file = "IC_iwatani.txt"
base_url = 'https://www.iwatani.co.jp'
access_url = 'https://www.iwatani.co.jp/jpn/news/'
target_url = ""
target_url1 = ""
target_url2 = ""
#############################################################################
#  list id変数定義(年に1回、毎年1月に新年度のデータが入ってきたら変数を更新する)
#  id_list2の変数にはid_list1の変数を変える。id_list1の変数に新しいid名を定義する
id_list1 = "pbBlock274889"
id_list2 = "pbBlock248469"
#############################################################################
max_row = 0
# base_file = "ICEnergyニュースリリース一覧テンプレート.xlsx"
export_file = "ICEnergyニュースリリース出力用" + date4 + ".xlsx"

# 企業名配列の定義
# company_array = []
# 決算変数の定義
output_company_name = ""

# write_flag = 0
xpath_str1 = ""

# 検索URL取得関数
def get_url(year):
    work_url = access_url + str(year) + "/"
    return work_url

# 年度でニュースリリースが変わるので、年度判定処理
def web_scrapping():
    w_ymd_array = today1.split('/')
    w_month = int(w_ymd_array[1])
    w_day = int(w_ymd_array[2])
    
    if w_month == 1 and w_day <= 12:
        target_url1 = get_url(date2)
        target_url2 = get_url(date3)
    else:
        target_url = get_url(date2)
    
    
    # Chromeを指定する
    driver = webdriver.Chrome()
    

    if w_month == 1 and w_day <= 12:
        for j in range(2):
            for url in [target_url1,target_url2]:
                # Chromeを開いて企業検索にアクセスする
                driver.get(url)
                sleep(3)
                # 年をまたいで30件のニュースリリースを取得 
                for i in range(1,16):
                    try:
                        if j == 0:
                            xpath_str1 = '//*[@id="' + id_list1 + '"]/div/table/tbody/tr[' + str(i) + ']/td[1]'
                            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                            print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
                            xpath_str2 = '//*[@id="' + id_list1 + '"]/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                            element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
                            print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
                        elif j == 1:
                            xpath_str1 = '//*[@id="' + id_list2 + '"]/div/table/tbody/tr[' + str(i) + ']/td[1]'
                            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                            print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
                            xpath_str2 = '//*[@id="' + id_list2 + '"]/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                            element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
                            print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))    
                    except:
                        if i == 1:
                            print("データがありません",file=codecs.open(file_name,'a','utf-8')) 
    else:
        driver.get(target_url)
        sleep(3)
        # 最新15件のニュースリリースを取得 
        for i in range(1,16):
            try:
                xpath_str1 = '//*[@id="pbBlock274889"]/div/table/tbody/tr[' + str(i) + ']/td[1]'
                element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
                print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
                xpath_str2 = '//*[@id="pbBlock274889"]/div/table/tbody/tr[' + str(i) + ']/td[2]/a'
                element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
                print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
            except:
                if i == 1:
                    print("データがありません",file=codecs.open(file_name,'a','utf-8'))
    # 画面を閉じる
    driver.quit()



def log_lotate():
    # 前回実行の作業ファイルがあれば削除
    file_exist = os.path.isfile(out_file)
    if file_exist:
        os.remove(out_file)

    # 作業用ログファイルをコピー
    shutil.copy2(file_name,out_file)

def get_max_row():
    w_row_count = 4
    while True:
        cell_value = ws.cell(row=w_row_count,column=3).value
        # print(cell_value)
        if cell_value == None:
            break
        else:
            w_row_count += 1
            # print(w_row_count)
    return w_row_count



web_scrapping()
company_name = "岩谷産業"
wb = op.load_workbook(export_file)
sh_name = '岩谷産業'
ws = wb[sh_name]
log_lotate()
max_row = get_max_row()
# ファイルを読み込み、URL、ニュースリリース日、ニュースリリース対象企業、リンクのタイトルを取得
fileobj = open(out_file,encoding="utf-8")
while True:
     line1 = fileobj.readline()
     line1 = line1.replace("\n","")
     if line1 == "":
         break
     result1 = re.match("								20",line1)
     result2 = re.match("<a href",line1)

     if result1:
         w_ymd = line1.replace("								","")
        #  print(w_ymd)


     if result2:
        w_array1 = line1.split(">")
        w_data1 = w_array1[0]
        w_array2 = w_data1.split("=")
        w_soutai_url = w_array2[1]
        w_soutai_url = w_soutai_url.replace('"','')
        w_soutai_url = w_soutai_url.replace(" class","")
        w_url = base_url + w_soutai_url    
        # print(w_url)
        w_title = w_array1[1]
        w_title = w_title.replace('<span class="icon_pdf"',"")
        w_title = w_title.replace('<span class="icon_new"',"")
        w_title = w_title.replace('</a',"")
        # print(w_title)
 
        ws.cell(row=max_row,column=3).value = company_name
        ws.cell(row=max_row,column=4).value = w_ymd
        ws.cell(row=max_row,column=5).value = w_title
        ws.cell(row=max_row,column=5).hyperlink = w_url
        ws.cell(row=max_row,column=6).value = today1
        ws.cell(row=max_row,column=7).value = w_url
                        
        max_row += 1
        # エクセルファイルの保存
        try:
           wb.save(export_file)
        except PermissionError as e:
           sys.exit()  

