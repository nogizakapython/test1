############ IC Energy 管理用エクセルファイルにニュースリリース日 ###########
############ 対象データを書き込む処理 ######################################
############ 新規作成  2024/3/13      ######################################
############ 作成者    takao.hattori  ######################################
############################################################################

# ライブラリのインポート
import openpyxl as op
from datetime import datetime
from openpyxl.styles.fonts import Font
from datetime import timedelta
# 日付の定義
dt = datetime.now()
date1 = dt.strftime('%Y%m%d')
# 管理ファイルの定義
admin_file = "SN_IC-Energy_石油関連企業のプレスリリース_日経News情報_20240312_v1.06.xlsx"
wb1 = op.load_workbook(admin_file)
sh_name = '管理'
ws1 = wb1[sh_name]
w1_row_count = 6
# 変数の定義
w_title = ""
w_url = ""
# 書き込み元のファイルの変数定義
input_file = "ICEnergyニュースリリース出力用" + date1 + ".xlsx"
wb2 = op.load_workbook(input_file)
sheet_array1 = ["コスモ石油","出光","エネオス","岩谷産業","岩谷産業IR","太陽石油","INPEX"]
# 入力日付の変数定義

# 前回の作業日を入力する関数
def data_input():
    #　前回の作業入力ルーチン
    
    ymd = ""
    while True:
        corect_count = 0
        print("前回の作業日付を西暦4桁、月2桁、日付2桁で入力してください。例) 2024年3月9日の場合は「20240309」と入力してください\n")
        data1 = input()
        if len(data1) == 8:
            corect_count += 1
        else:
            print("入力文字数が違います。")
        
        w_year = data1[:4]
        w_month = data1[4:6]
        w_day = data1[6:]
        if int(w_month) < 1 or int(w_month) > 12:
            print("NG")
        else:     
            corect_count += 1
    
        if int(w_day) < 1 or int(w_day) > 31:
            print("NG")
        else:     
            corect_count += 1
        if corect_count == 3:
            ymd = w_year + '/' + w_month + '/' + w_day
            break 
    return ymd        

# 管理用エクセルファイルの書き込み開始行を取得する関数
def admin_excel_maxcount():
    admin_row_count = 6
    while True:
        value1 = ws1.cell(row=admin_row_count,column=3).value
        if value1 == None:
            break
        else:
            admin_row_count += 1
            # print(admin_row_count)
    return admin_row_count    

# 企業別前回のニュースリリース日を先週分の最終リリース日のセルに書き込む関数
def admin_before_workday():
    if sh_name == "コスモ石油":
        ws1.cell(row=2,column=4).value = ws1.cell(row=3,column=4).value
    if sh_name == "出光":
        ws1.cell(row=2,column=5).value = ws1.cell(row=3,column=5).value
    if sh_name == "エネオス":
        ws1.cell(row=2,column=6).value = ws1.cell(row=3,column=6).value
    if sh_name == "岩谷産業":
        ws1.cell(row=2,column=7).value = ws1.cell(row=3,column=7).value
    if sh_name == "岩谷産業IR":
        ws1.cell(row=2,column=8).value = ws1.cell(row=3,column=7).value    
    if sh_name == "太陽石油":
        ws1.cell(row=2,column=9).value = ws1.cell(row=3,column=8).value
    if sh_name == "INPEX":
        ws1.cell(row=2,column=10).value = ws1.cell(row=3,column=9).value
    wb1.save(admin_file)

# 最終ニュースリリース日をセルに書き込む関数
def admin_now_workday(now_workday):
    if sh_name == "コスモ石油":
        ws1.cell(row=3,column=4).value = now_workday
    if sh_name == "出光":
        ws1.cell(row=3,column=5).value = now_workday    
    if sh_name == "エネオス":
        ws1.cell(row=3,column=6).value = now_workday
    if sh_name == "岩谷産業":
        ws1.cell(row=3,column=7).value = now_workday
    if sh_name == "岩谷産業IR":
        ws1.cell(row=3,column=8).value = now_workday    
    if sh_name == "太陽石油":
        ws1.cell(row=3,column=9).value = now_workday
    if sh_name == "INPEX":
        ws1.cell(row=3,column=10).value = now_workday 

# メイン処理
str_ymd = data_input()

# 今回の作業日と前回の作業日の日付差を取得し、前回の作業日から今回の作業日の前日まで日付を
# 配列に格納する。
# print(str_ymd)
ymd = datetime.strptime(str_ymd, '%Y/%m/%d')
array1 = []
dt = datetime.now()
day_sabun = dt - ymd
day_sabun = str(day_sabun)
sabun_array1 = day_sabun.split(" ")
day_sa = sabun_array1[0]
day_sa = int(day_sa)
# 前回の作業日から今回の作業前日までを日付配列に格納する。
for i in range(day_sa):
    w_ymd = ymd.strftime("%Y/%m/%d")
    array1.append(w_ymd)
    ymd = ymd + timedelta(1)

# 管理用エクセルファイルの書き込み開始行を取得する
w1_row_count = admin_excel_maxcount()

# 石油各社6社別シートからスクレイピングデータを読み込む 
for sh_name in list(sheet_array1):
    ws2 = wb2[sh_name]
    w2_row_count = 5
    admin_before_workday()
    # 各企業別のエクセルシートの最終行を取得する
    while True:
        data_value = ws2.cell(row=w2_row_count,column=5).value
        if data_value == None:
            break
        else:
            w2_row_count += 1
    #管理用シートに書き込むデータをニュースリリース日と前回の作業日から昨日までの日付で
    #突き合わせして、合致したものを管理用エクセルファイルに書き込む  
    for row_count in range(w2_row_count-1,4,-1):
        ad_company = ws2.cell(row=row_count,column=3).value
        ad_ymd = ws2.cell(row=row_count,column=4).value
        ad_title = ws2.cell(row=row_count,column=5).value
        ad_workday = ws2.cell(row=row_count,column=6).value
        ad_url = ws2.cell(row=row_count,column=7).value
        for t_ymd in array1:
            if ad_ymd == t_ymd:
                ws1.cell(row=w1_row_count,column=3).value = ad_company
                ws1.cell(row=w1_row_count,column=4).value = ad_ymd
                ws1.cell(row=w1_row_count,column=5).value = ad_title
                ws1.cell(row=w1_row_count,column=5).hyperlink = ad_url
                ws1.cell(row=w1_row_count,column=5).font = Font(color='0000FF',underline='single',name='Meiryo UI')
                ws1.cell(row=w1_row_count,column=6).hyperlink = ad_workday
                admin_now_workday(ad_ymd)
                w1_row_count += 1
                wb1.save(admin_file)    
    
    