############ テスト  #################
######################################
import openpyxl as op
from datetime import datetime
from openpyxl.styles.fonts import Font

dt = datetime.now()
date1 = dt.strftime('%Y%m%d')
# 管理ファイルの定義
admin_file = "SN_IC-Energy_石油関連企業のプレスリリース_日経News情報_20240312_v1.06.xlsx"
wb1 = op.load_workbook(admin_file)
sh_name = '管理'
ws1 = wb1[sh_name]
w1_row_count = 6
w_title = ""
w_url = ""
# 書き込み元のファイル
input_file = "ICEnergyニュースリリース出力用" + date1 + ".xlsx"
wb2 = op.load_workbook(input_file)
sheet_array1 = ["コスモ石油","出光","エネオス","岩谷産業","太陽石油","INPEX"]


while True:
    value1 = ws1.cell(row=w1_row_count,column=3).value
    if value1 == None:
        break
    else:
        w1_row_count += 1



for sh_name in list(sheet_array1):
    ws2 = wb2[sh_name]
    w2_row_count = 5
    
    while True:
        data_value = ws2.cell(row=w2_row_count,column=5).value
        if data_value == None:
            break
        else:
            w2_row_count += 1
            
    for row_count in range(w2_row_count-1,4,-1):
        
        w_ymd = ""
        for i in range(3,8):
            if i == 4:
                w_ymd = ws2.cell(row=row_count,column=i).value
                 
            if i == 5:
                w_title = ws2.cell(row=row_count,column=i).value
            elif i == 7:
                w_url = ws2.cell(row=row_count,column=i).value
                ws1.cell(row=w1_row_count,column=5).value = w_title
                ws1.cell(row=w1_row_count,column=5).hyperlink = w_url
                ws1.cell(row=w1_row_count,column=5).font = Font(color='0000FF',underline='single')
            else:    
                ws1.cell(row=w1_row_count,column=i).value = ws2.cell(row=row_count,column=i).value      
        w1_row_count += 1
    wb1.save(admin_file)
    if row_count == 5:
        if sh_name == "コスモ石油":
            ws1.cell(row=2,column=4).value = ws1.cell(row=3,column=4).value
            ws1.cell(row=3,column=4).value = w_ymd
        if sh_name == "出光":
            ws1.cell(row=2,column=5).value = ws1.cell(row=3,column=5).value
            ws1.cell(row=3,column=5).value = w_ymd    
        if sh_name == "エネオス":
            ws1.cell(row=2,column=6).value = ws1.cell(row=3,column=6).value
            ws1.cell(row=3,column=6).value = w_ymd
        if sh_name == "岩谷産業":
            ws1.cell(row=2,column=7).value = ws1.cell(row=3,column=7).value
            ws1.cell(row=3,column=7).value = w_ymd
        if sh_name == "太陽石油":
            ws1.cell(row=2,column=8).value = ws1.cell(row=3,column=8).value
            ws1.cell(row=3,column=8).value = w_ymd
        if sh_name == "INPEX":
            ws1.cell(row=2,column=9).value = ws1.cell(row=3,column=9).value
            ws1.cell(row=3,column=9).value = w_ymd 
    wb1.save(admin_file)      

