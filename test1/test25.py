#######   myTIM勤務終了ボタン処理　###########
#######   新規作成  2023/11/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium

dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')


file_name = "shane" + date1 + ".txt"
out_file = "shane.txt"
string1 = '<div id="sideCallouts">'
date_str = ""
w_title = ""
base_url = 'https://www.shane.co.jp'

target_url = 'https://www.shane.co.jp/news/'
max_row = 5
#base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
# Chromeを指定する

driver = webdriver.Chrome()

# Chromeを開いてMyTIMにアクセスする
try:
    driver.get(target_url)
    sleep(5)
    
    
    

        
    for i in range(1,31):
        
        
        xpath_str1 = '//*[@id="main_module"]/section[2]/div/ul/li[' + str(i) + ']/a'
        try:
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        except:
            break
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        xpath_str2 = '//*[@id="main_module"]/section[2]/div/ul/li[' + str(i) + ']/a/p[1]'
        element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
        print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        xpath_str3 = '//*[@id="main_module"]/section[2]/div/ul/li[' + str(i) + ']/a/p[2]'
        element_str3 = driver.find_element(by=By.XPATH,value=xpath_str3)
        print(element_str3.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))

        

        

except EnvironmentError as e:
    str100 = e     


# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(file_name,out_file)



fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result1 = re.match("<a href",line1)
    result5 = re.match("<a rel",line1)
    result2 = re.match('<p class="newsList01_date"',line1)
    result3 = re.match('<p class="newsList01_txt"',line1)


    if result1:
        w_array1 = line1.split(" ")
        w_url = w_array1[1]
        w_url = w_url.replace('href=',"")
        w_url = w_url.replace('"',"")
        #print(w_url)

    if result5:
        w_array5 = line1.split(" ")
        w_url = w_array5[2]
        w_url = w_url.replace('href=',"")
        w_url = w_url.replace('"',"")
        #print(w_url)
 


    if result2:
        w_array2 = line1.split(">")
        w_line = w_array2[1]
        w_line = w_line.replace('</p',"")
        w_line = w_line.replace('.',"/")
        w_ymd = w_line
        #print(w_ymd)  

    if result3:
        w_array3 = line1.split(">")
        w_title = w_array3[1]
        w_title = w_title.replace("</p","")
        #print(w_title)
        key_word = key_word = r"(役員|異動)"
        result4 = re.search(key_word,w_title)
        if result4:
            wb = op.load_workbook(export_file)
            sh_name = 'SHANE CORPORATION'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)     

