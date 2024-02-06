#######   myTIM勤務終了ボタン処理　###########
#######   新規作成  2023/11/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
import openpyxl as op
import codecs

dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')

file_name = "test" + date1 + ".txt"
out_file = "bausch.txt"
string1 = '<div id="sideCallouts">'
date_str = ""
w_title = ""
base_url = 'https://www.bausch.co.jp'
target_url = 'https://www.bausch.co.jp/ja-jp/our-company/newsroom/' + date2 + '/' 
max_row = 5
#base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0

# Chromeを指定する
driver = webdriver.Chrome()
# Chromeを開いてMyTIMにアクセスする
try:
    driver.get(target_url)
except EnvironmentError as e:
    str100 = e     


html = driver.page_source


try:
    print(html,file=codecs.open(file_name,'a','utf-8'))
        
    

    # ファイルへの書き込みエラー時、例外処理を実行し、エラーをコンソールに出力する。
except IOError as e:
        print("Do not Write Result File!")

        print(e)
        exit
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)


with open(file_name,encoding="utf-8") as f:
    lines = f.readlines()
    for line in lines:
        line = line.replace("\n","")
        if line == '	<div id="contentHolder">':
            write_flag = 1

        result1 = re.search("<h6>",line)
        result2 = re.match("                    <a href=",line)
        result3 = re.match('                        ..',line)
        result4 = re.match(string1,line)
        if write_flag == 1:
            if result1:
                try:
                    print(line,file=codecs.open(out_file,'a','utf-8'))
                except IOError as e:
                    print("Do not Write Result File!")
                    print(e)
                    exit
            if result2:
                try:
                    print(line,file=codecs.open(out_file,'a','utf-8'))
                except IOError as e:
                    print("Do not Write Result File!")
                    print(e)
                    exit
            if result3:
                try:
                    print(line,file=codecs.open(out_file,'a','utf-8'))
                except IOError as e:
                    print("Do not Write Result File!")
                    print(e)
                    exit        
            if result4:
                break

fileobj = open(out_file,encoding="utf-8")


while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result4 = re.search("<h6>",line1)
    result5 = re.search("<a href",line1)
    result6 = re.match("                        ",line1)
    if result4:
        line1 = line1.replace("               <h6>","")
        line1 = line1.replace("</h6>","")
        date_str = line1
        w_array1 = date_str.split(',')
        result_date = w_array1[0]
        #print(date_str)
            
    if result5:
        w_array1 = line1.split("=")
        w_url = w_array1[1]
        w_url = w_url.replace('"','')
        w_url = w_url.replace(">","")
        w_url = base_url + w_url

    if result6:
        w_title = line1.replace(" ","")    
        key_word = r"(役員|異動)"
        result7 = re.search(key_word,w_title)
        if result7:
                    
            wb = op.load_workbook(export_file)
            sh_name = 'ボシュロム・ジャパン'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = result_date
            ws.cell(row=max_row,column=6).value = w_url
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)        
                                    

                



            







