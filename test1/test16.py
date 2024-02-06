#######   myTIM勤務終了ボタン処理　###########
#######   新規作成  2023/11/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium

dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')


file_name = "koreanair" + date1 + ".txt"
out_file = "koreanair.txt"
string1 = '<div id="sideCallouts">'
date_str = ""
w_title = ""
base_url = 'https://www.koreanair.com'

target_url = 'https://www.koreanair.com/jp/ja/footer/about-us/newsroom/list'
max_row = 5
#base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
# Chromeを指定する

driver = webdriver.Chrome()

# Chromeを開いてMyTIMにアクセスする
try:
    driver.get(target_url)
    sleep(5)
    
    
    for i in range(1,30):
        xpath_str1 = '/html/body/div[9]/div/div/div/div/ul/li[' + str(i) + ']/div[2]/a'
        try:
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        except:
            break
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        
        xpath_str2 = '/html/body/div[9]/div/div/div/div/ul/li[' + str(i) + ']/div[2]/a/strong'
        element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
        print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        xpath_str3 = '/html/body/div[9]/div/div/div/div/ul/li[' + str(i) + ']/div[2]/a/span'
        element_str3 = driver.find_element(by=By.XPATH,value=xpath_str3)
    

except EnvironmentError as e:
    str100 = e     


# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(file_name,out_file)



fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result1 = re.match("<a href",line1)
    result2 = re.search("<span",line1)
    result3 = re.match("<strong",line1)
    if result1:
        w_array1 = line1.split("=")
        w_url = w_array1[1]
        w_url = w_url.replace('>',"")
        w_url = w_url.replace('"',"")
        w_url = base_url + w_url
        #print(w_url)

    if result2:
        w_array1 = line1.split(">")
        w_line = w_array1[1]
        w_line = w_line.replace(" ","")
        w_line = w_line.replace("年","/")
        w_line = w_line.replace("月","/")
        w_line = w_line.replace("日","")
        w_line = w_line.replace("</span","")
        w_ymd = w_line       
        #print(w_ymd)   
    
    if result3:
        w_array3 = line1.split(">")  
        w_title = w_array3[1]
        w_title = w_title.replace("</strong","")
        #print(w_title)
        key_word = key_word = r"(人事|異動|就任)"
        result4 = re.search(key_word,w_title)
        if result4:
            wb = op.load_workbook(export_file)
            sh_name = 'KOREAN AIR'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)        
         
