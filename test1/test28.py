#######   myTIM勤務終了ボタン処理　###########
#######   新規作成  2023/11/27  ##########
#######   Author  takao.hattori ###########


# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
import openpyxl as op
import codecs

dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')

file_name = "tokan" + date1 + ".txt"
out_file = "tokan.txt"
string1 = '<div id="sideCallouts">'
date_str = ""
w_title = ""
base_url = 'https://www.tokan.co.jp'
target_url = 'https://www.tokan.co.jp/news/' 
max_row = 5
#base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
w_ymd = ""
w_tilte = ""
w_url = ""
w_url_flag = 0

# Chromeを指定する
driver = webdriver.Chrome()
# Chromeを開いてMyTIMにアクセスする
try:
    driver.get(target_url)
except EnvironmentError as e:
    str100 = e     


html = driver.page_source


try:
    print(html,file=codecs.open(file_name,'a','utf-8'))
        
    

    # ファイルへの書き込みエラー時、例外処理を実行し、エラーをコンソールに出力する。
except IOError as e:
        print("Do not Write Result File!")

        print(e)
        exit
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)


with open(file_name,encoding="utf-8") as f:
    lines = f.readlines()
    for line in lines:
        line = line.replace("\n","")

        result1 = re.match('                    <h2>ニュース・新着情報</h2>',line)
        result2 = re.match("    </main>",line)
        
        result6 = re.search('pdf',line)
        
        if result2:
            write_flag = 2
        if result1:
            write_flag = 1
            
        if write_flag == 1:    
            print(line,file=codecs.open(out_file,'a','utf-8'))

with open(out_file,encoding="utf-8") as f2:
    lines = f2.readlines()
    for line1 in lines:
        line1 = line1.replace("\n","")
        result3 = re.search("<time>",line1)
        result4 = re.match("                            <h3>",line1)
        result5 = re.search(f'人事|役員',line1)
        result6 = re.match("                                        <a href=",line1)
        if result3:
            w_array1 = line1.split(">")
            w_line = w_array1[1]
            w_ymd = w_line.replace("</time","")
            w_ymd = w_ymd.replace(".","/")
            #print(w_ymd)
        if result4:
            w_array2 = line1.split(">")
            w_title = w_array2[1]
            w_title = w_title.replace("</h3","")
            #print(w_title)
            if result5:
                w_url_flag = 1
        if w_url_flag == 1 and result6:
            w_array3 = line1.split("<")
            w_url = w_array3[1]
            w_url = w_url.replace("a href=","")
            w_url = w_url.replace('"',"")
            w_url = w_url.replace(" target=_blank class=pdf>","")
            print(w_url)
            wb = op.load_workbook(export_file)
            sh_name = 'TOKAN CO. LTD'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
                
            max_row += 1
            # エクセルファイルの保存
            wb.save(export_file)
            w_url_flag = 0
