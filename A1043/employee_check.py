###############################################
####    社員ID、社員名突合処理確認処理　　　#####
####    新規作成日:   2024/6/18            #####
################################################

# 突合ファイル削除関数
def remove_file(file_name):
    import os
    file_check = os.path.isfile(file_name)
    if file_check:
        os.remove(file_name)

# 配列初期化関数 
def initial_array():
    array1 = []
    return array1

# 配列の要素の一意化関数
def unique_make_array(array1):
    w_array = list(set(array1))
    return w_array        

# 突合結果出力関数
def output_result_file(msg1,file_name):
    import codecs
    print(msg1,file=codecs.open(file_name,'a','utf-8'))


# メイン関数
def main():
    # ライブラリのインポート
    import openpyxl as op

    # 社員テーブルexportファイルの変数定義
    read_file1 = "社員.xlsx"
    wb1 = op.load_workbook(read_file1)
    sh_name1 = '社員'
    ws1 = wb1[sh_name1]
    w1_row_count = 2

    # 社員テーブルexportファイルの変数定義
    read_file2 = "ファイルエクスポート用.xlsm"
    wb2 = op.load_workbook(read_file2)
    sh_name2 = 'Sheet1'
    ws2 = wb2[sh_name2]
    w2_row_count = 2
    
    # 突合結果ファイル
    id_not_file="result1.txt"

    # 前回作業用突合結果ファイルの削除
    remove_file(id_not_file)

    # ACN社員の日報から作業者の社員番号を取得し、社員番号を一意にする。
    while True:
        if w2_row_count == 2:
            w_array = initial_array()

        employee_id = ws2.cell(row=w2_row_count,column=2).value
        if employee_id == None:
            break
        else:
            w_array.append(str(employee_id))
        w2_row_count += 1
    
    report_id_array = unique_make_array(w_array) 

    # テーブルからエクスポートした社員マスタファイルから、社員番号を配列に格納し、一意にする。
    while True:
        if w1_row_count == 2:
            w_array = initial_array()

        export_id = ws1.cell(row=w1_row_count,column=2).value
        
        if export_id == None:
            break
        else:
            w_array.append(export_id)
        w1_row_count += 1   
            
    table_array = list(set(w_array))

    # 突合処理
    for employee_id in report_id_array:
        correct_count = 0
        length1 = len(table_array)
        for i in range(0,length1):
            t_employee_id = table_array[i]
            if employee_id == t_employee_id:
                correct_count += 1
            if i == length1 - 1:
                if correct_count == 1:
                    output_result_file(f"{employee_id}は社員テーブルに存在します",id_not_file)
                else:
                    output_result_file(f"{employee_id}は社員テーブルに存在しません",id_not_file)

if __name__ == "__main__":
    main()
