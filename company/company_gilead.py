#######   ギリアドサイエンス　企業HPから人事情報を取得する　###########
#######   新規作成  2023/11/27  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys 
import company_fileopenerror

dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')


file_name = "gilead" + date1 + ".txt"
out_file = "gilead.txt"
string1 = '<div id="sideCallouts">'
date_str = ""
w_title = ""
# 2025/10/27 baseURL変更に伴う修正 takao.hattori
base_url = 'https://www.gilead.com'
# 2025/10/27 URL変更に伴う修正
target_url = 'https://www.gilead.com/ja-jp/news'
max_row = 5
#base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
# Chromeを指定する

driver = webdriver.Chrome()
file1 = open(file_name,"w")
file1.write("")
file1.close()



# Chromeを開いて企業HPのプレスリリースにアクセスする
try:
    driver.get(target_url)
    sleep(5)
    # 取得件数を500→100件に減らす対応 2025/5/2 takao.hattori
    for i in range(1,101):
        # xpathの修正 2025/5/2 takao.hattori
        # xpathの修正 2025/10/27 takao.hattori
        xpath_str1 = '/html/body/div[1]/main/section[3]/div/div/div/div[1]/ul/li[' + str(i) + ']/a'
        

                       
        try:
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        except:
            break
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        # xpathの修正 2025/5/2 takao.hattori
        # xpath_str2 = '/html/body/div[1]/main/div/section/div[' + str(i) + ']/a'
        # element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
        # print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))    
    

except EnvironmentError as e:
    str100 = e     


# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(file_name,out_file)



fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    # 2025/10/27 HTML変更に伴うタグ検索条件の修正
    result1 = re.search("gl-notes-date",line1)
    result2 = re.search("<h5>",line1)
    result3 = re.match('<a target="_blank"',line1)
    # 2025/10/27 HTML掲載日付タグ変更に伴う修正
    # 2026/6/1   HTML掲載日付タグ変更に伴う修正
    if result1:
        w_array1 = line1.split(">")
        w_y = w_array1[1]
        w_y = w_y.replace("<strong","")
        w_y = w_y.replace(".","/")
        w_y = w_y.replace("年","/")
        w_y = w_y.replace("月","/")
        w_y = w_y.replace("日","/")
        w_array2 = w_y.split("/")
        w_y = w_array2[0]
        w_m =  w_array2[1]
        w_d = w_array2[2]
       
        w_m2 = int(w_m)
        if w_m2 < 10:
            w_m = "0" + w_m
        
        w_d2 = int(w_d)
        if w_d2 < 10:
            w_d = "0" + w_d
        w_ymd = w_y + "/" + w_m + "/" + w_d       
        # print(w_ymd)
    # 2025/10/27 HTML表題のタグ変更に伴う修正   
    if result2:
        w_array3 = line1.split(">")
        w_title = w_array3[1]
        w_title = w_title.replace("</h5","")
        # print(w_title)
    # 2025/10/27 HTML掲載先のURL変更に伴う修正    
    if result3:
        w_array4 = line1.split("=")
        w_url = w_array4[3]
        w_url = w_url.replace('">',"")
        w_url = w_url.replace('"',"")
        w_url = base_url + w_url
        # print(w_url)
        
        key_word = key_word = r"(就任|任命|交代)"
        reject_word = "アンバサダー"
        result4 = re.search(key_word,w_title)
        result5 = re.search(reject_word,w_title)
        if result4:
            if result5:
                continue
            else:
                wb = op.load_workbook(export_file)
                sh_name = 'ギリアド・サイエンシズ'
                ws = wb[sh_name]
                ws.cell(row=max_row,column=2).value = w_title
                ws.cell(row=max_row,column=3).value = w_url
                ws.cell(row=max_row,column=4).value = w_ymd
                ws.cell(row=max_row,column=6).value = w_url
                ws.cell(row=max_row,column=6).hyperlink = w_url
                
                max_row += 1
#                 # エクセルファイルの保存
                try:
                    wb.save(export_file)
                except PermissionError as e:
                    fname = export_file
                    openerr = company_fileopenerror.ReadfileError(fname)
                    openerr.readerror()
                    sys.exit()               
         
