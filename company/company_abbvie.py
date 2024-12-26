#######   アッビィジャパン　企業HPから人事情報を取得する　###########
#######   新規作成  2023/12/26  ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
import openpyxl as op
import shutil
import sys
import codecs
import company_fileopenerror


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')

file_name = "abbvie" + date1 + ".txt"
out_file = "abbvie.txt"
string1 = "    <div class=\"separator\">"
date_str = ""
w_title = ""
base_url = 'https://www.abbvie.co.jp'
target_url = 'https://www.abbvie.co.jp/press-release/' + date2 + '-news-archive.html'
max_row = 5
base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
length1 = 0

# Chromeを指定する
driver = webdriver.Chrome()
# Chromeを開いてMyTIMにアクセスする
driver.get(target_url)


try:
    shutil.copy2(base_file,export_file)
except FileNotFoundError:
    print("リネーム前のファイルが存在しません")
    sys.exit()
except PermissionError as e:
    fname = export_file
    openerr = company_fileopenerror.ReadfileError(fname)
    openerr.readerror()
    sys.exit()
                                    

html = driver.page_source


try:
    print(html,file=codecs.open(file_name,'a','utf-8'))
        
    

    # ファイルへの書き込みエラー時、例外処理を実行し、エラーをコンソールに出力する。
except IOError as e:
        print("Do not Write Result File!")

        print(e)
        exit
# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
    os.remove(out_file)


with open(file_name,encoding="utf-8") as f:
    lines = f.readlines()
    for line in lines:
        line = line.replace("\n","")
        result1 = re.search("<h3>",line)
        result2 = re.match("<p><a href",line)
        result3 = re.match(string1,line)
        if result1:
            try:
                print(line,file=codecs.open(out_file,'a','utf-8'))
            except IOError as e:
                print("Do not Write Result File!")
                print(e)
                exit
        if result2:
            try:
                print(line,file=codecs.open(out_file,'a','utf-8'))
            except IOError as e:
                print("Do not Write Result File!")
                print(e)
                exit
        if result3:
            break

fileobj = open(out_file,encoding="utf-8")


while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break

    # 2024/12/26 HPデザイン変更に伴う修正  
    result4 = re.match("    <h3>",line1)
    result5 = re.match("<div",line1)
    
    
    #　タイトルとURLの取得条件が「>」区切りで配列に修正
    # 2024/12/26 takao.hattori 
    if result5:
        w_array1 = line1.split(">")
        length1 = len(w_array1)
        # 配列の要素の個数が8より大きい場合、配列にして、日付とPDFのURLを取得する
        if length1 > 8:
             w_url = w_array1[8]
             url_array = w_url.split(" ")
             w_url = url_array[1]
             w_url = w_url.replace("href=\&quot;","")
             w_url = w_url.replace("\&quot;","")
             w_url = base_url + w_url
             w_url = w_url.replace('"','')
            #  print(w_url)
             w_title = w_array1[9]
             w_title = w_title.replace("</a","")
            #  print(w_title)

    if result4:
        array1 = line1.split(">")
        date_str = array1[3]
        date_str = date_str.replace("</b","")
        date_str = date_str.replace("&nbsp;"," ")
        # print(date_str)
        if  length1 > 8: 
            key_word = r"(人事|異動)"
            result6 = re.search(key_word,w_title)
            if result6:
                 w_array3 = date_str.split(",")
                 date_array = w_array3[0].split(" ")
                 date_y = w_array3[1]
                 date_y = date_y.replace(" ","")
                 date_m = date_array[0]
                 date_d = date_array[1]
                 date_m2 = ""
                 if date_m == "Jan":
                     date_m2 = "01"
                 elif date_m == "Feb":
                     date_m2 = "02"
                 elif date_m == "Mar":
                     date_m2 = "03"
                 elif date_m == "Apr":
                     date_m2 = "04"
                 elif date_m == "May":
                     date_m2 = "05"
                 elif date_m == "Jun":
                     date_m2 = "06"
                 elif date_m == "Jul":
                     date_m2 = "07"
                 elif date_m == "Aug":
                     date_m2 = "08"
                 elif date_m == "Sep":
                     date_m2 = "09"
                 elif date_m == "Oct":
                     date_m2 = "10"
                 elif date_m == "Nov":
                     date_m2 = "11"
                 elif date_m == "Dec":
                     date_m2 = "12"

                 date_d2 = int(date_d)
                 if date_d2 < 10:
                    date_d2 = "0" + str(date_d2) 
                 else:
                    date_d2 = str(date_d2)      
                 w_date =  str(date_y) + "/" + date_m2 + "/" + str(date_d2)
                 wb = op.load_workbook(export_file)
                 sh_name = 'アッヴィ'
                 ws = wb[sh_name]
                 ws.cell(row=max_row,column=2).value = w_title
                 ws.cell(row=max_row,column=3).value = w_url
                 ws.cell(row=max_row,column=4).value = w_date
                 ws.cell(row=max_row,column=6).value = w_url
                 ws.cell(row=max_row,column=6).hyperlink = w_url
                           
                 max_row += 1
                 wb.save(export_file)
            
