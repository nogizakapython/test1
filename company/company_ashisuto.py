#######   アシスト 企業HP勤務終了ボタン処理　###########
#######   新規作成  2023/12/28  ##########
#######   webサイトリニューアルによる修正  2025/7/25  takao.hattori ##########
#######   Author  takao.hattori ###########



# 時間を計るライブラリをインポート
import datetime
import re
import os
# WebDriverライブラリをインポート
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as op
import codecs
from time import sleep
import shutil
import selenium
import sys
import company_fileopenerror


dt = datetime.datetime.now()
date1 = dt.strftime('%Y%m%d%H%M%S')
date2 = dt.strftime('%Y')
date3 = dt.strftime('%Y%m%d')


file_name = "ashisuto" + date1 + ".txt"
out_file = "ashisuto.txt"
string1 = '<div id="sideCallouts">'
date_str = ""
w_title = ""
base_url = 'https://www.ashisuto.co.jp'

target_url = 'https://www.ashisuto.co.jp/news/' + date2 + ".html"
max_row = 5
#base_file = "【企業個別】検索結果_yyyymmdd.xlsx"
export_file = "【企業個別】検索結果_" + date3 + ".xlsx"
row_count = 0
write_flag = 0
xpath_str1 = ""
# Chromeを指定する

driver = webdriver.Chrome()

# Chromeを開いて企業HPにアクセスする
try:
    driver.get(target_url)
    sleep(5)
    
    for i in range(1,70):
    
    #    xpathの変更 2025/7/25 takao.hattori
        xpath_str1 = '/html/body/div[1]/main/div/article/div[2]/div[2]/ul/li[' + str(i) + ']/time/span[1]'
        

        try:
            element_str1 = driver.find_element(by=By.XPATH,value=xpath_str1)
        except:
            break
        print(element_str1.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        xpath_str2 = '/html/body/div[1]/main/div/article/div[2]/div[2]/ul/li[' + str(i) + ']/div/a'
        
        element_str2 = driver.find_element(by=By.XPATH,value=xpath_str2)
        print(element_str2.get_attribute("outerHTML"),file=codecs.open(file_name,'a','utf-8'))
        
    

except EnvironmentError as e:
    str100 = e     


# 画面を閉じる
driver.quit()
file_exist = os.path.isfile(out_file)
if file_exist:
   os.remove(out_file)

shutil.copy2(file_name,out_file)



fileobj = open(out_file,encoding="utf-8")
while True:
    line1 = fileobj.readline()
    line1 = line1.replace("\n","")
    if line1:
        row_count += 1
    else:
        break   
    result1 = re.match("<span",line1)
    result2 = re.match("<a href",line1)

    if result1:
        w_array1 = line1.split(">")
        w_line = w_array1[1]
        w_array2 = w_line.split("(")
        w_line = w_array2[0]
        w_line = w_line.replace("年","/")
        w_line = w_line.replace("月","/")
        w_line = w_line.replace("日","")
        w_line = w_line.replace("</span","")
        w_ymd = w_line       
        # print(w_ymd)   
    
    if result2:
        w_array2 = line1.split(">")
        w_url = w_array2[0]
        w_url = w_url.replace('<a href=',"")
        w_url = w_url.replace('"',"")
        w_url = base_url + w_url
        # print(w_url)
        w_title = w_array2[1]
        w_title = w_title.replace("</a","")
        # print(w_title)
        key_word = key_word = r"(役員)"
        result3 = re.search(key_word,w_title)
        if result3:
            wb = op.load_workbook(export_file)
            sh_name = 'アシスト'
            ws = wb[sh_name]
            ws.cell(row=max_row,column=2).value = w_title
            ws.cell(row=max_row,column=3).value = w_url
            ws.cell(row=max_row,column=4).value = w_ymd
            ws.cell(row=max_row,column=6).value = w_url
            ws.cell(row=max_row,column=6).hyperlink = w_url
                
            max_row += 1
            # エクセルファイルの保存
            try:
                wb.save(export_file)
            except PermissionError as e:
                fname = export_file
                openerr = company_fileopenerror.ReadfileError(fname)
                openerr.readerror()
                sys.exit()             
         
