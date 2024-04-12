import openpyxl as op
import random

input_file = "drawing.xlsx"
sheet_name = "data"
array_list = []

start_num = 3
eid_count = start_num
end_num = 0

wb = op.load_workbook(input_file)
ws = wb[sheet_name]

while True:
    eid = ws.cell(row=eid_count,column=2).value
    status = ws.cell(row=eid_count,column=3).value

    if eid == None:
        break
    else:
        if status == "〇":
            array_list.append(eid)
        
        eid_count += 1
        
 

random.shuffle(array_list)

while True:
    print("レスポンスなうに割り当てる人数を1～19の整数で入力してください")
    try:
        num = int(input())
        if num > 0 and num < 20:
            break
        else:
            print("1以上19以下の正の整数を入力してください")
    except ValueError:
        print("文字列を入力しないでください")


for i in range(0,num):
    name = array_list[i]
    print(name)



